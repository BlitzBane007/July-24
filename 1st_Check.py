import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import messagebox
from tqdm import tqdm


def read_csv_file(filename, delimiter=','):
    rows = []
    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            rows.append(row)
    return rows


def read_xlsx_file(filename):
    rows = []
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows


def find_missing_headers(file1_headers, file2_headers):
    missing_file1 = set(file2_headers) - set(file1_headers)
    missing_file2 = set(file1_headers) - set(file2_headers)
    return missing_file1, missing_file2


def find_column_index(header_row, column_name):
    return header_row.index(column_name) if column_name in header_row else -1


def find_missing_data(file1_data, file2_data, column_index1, column_index2):
    file1_values = set(row[column_index1] for row in tqdm(file1_data[1:], desc='Processing File 1', unit='row',
                                                          ncols=80)) if column_index1 >= 0 else set()
    file2_values = set(row[column_index2] for row in tqdm(file2_data[1:], desc='Processing File 2', unit='row',
                                                          ncols=80)) if column_index2 >= 0 else set()
    missing_file1 = file2_values - file1_values
    missing_file2 = file1_values - file2_values
    return missing_file1, missing_file2


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[('CSV Files', '*.csv'), ('Excel Files', '*.xlsx')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def browse_output_file(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[('CSV Files', '*.csv')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def process_files():
    # Get file paths
    file_path_1 = entry_file1.get()
    file_path_2 = entry_file2.get()

    # Get delimiter and header indices
    delimiter1 = entry_delimiter1.get()
    delimiter2 = entry_delimiter2.get()
    header_index_1 = int(entry_header1.get())
    header_index_2 = int(entry_header2.get())

    # Check if files are selected
    if not file_path_1 or not file_path_2:
        messagebox.showerror("Error", "Please select both files.")
        return

    # Read files based on format
    if file_path_1.lower().endswith('.csv'):
        file1_data = read_csv_file(file_path_1, delimiter1)
    elif file_path_1.lower().endswith('.xlsx'):
        file1_data = read_xlsx_file(file_path_1)
    else:
        messagebox.showerror("Error", "Unsupported file format for File 1.")
        return

    if file_path_2.lower().endswith('.csv'):
        file2_data = read_csv_file(file_path_2, delimiter2)
    elif file_path_2.lower().endswith('.xlsx'):
        file2_data = read_xlsx_file(file_path_2)
    else:
        messagebox.showerror("Error", "Unsupported file format for File 2.")
        return

    # Rest of the code for comparison and generating output file...
    file1_headers = file1_data[header_index_1]
    file2_headers = file2_data[header_index_2]
    header1count = 0
    header2count = 0
    for header1 in file1_headers:
        header1count = header1count+1
    print(header1count)
    for header2 in file2_headers:
        header2count = header2count+1
    print(header2count)


    missing_file1, missing_file2 = find_missing_headers(file1_headers, file2_headers)

    column_name = 'OFST020000'
    column_index_file1 = find_column_index(file1_headers, column_name)
    column_index_file2 = find_column_index(file2_headers, column_name)

    missing_file1_data, missing_file2_data = find_missing_data(file1_data, file2_data, column_index_file1,
                                                               column_index_file2)

    output_path = entry_output_file.get()
    with open(output_path, 'w', encoding='utf-8', newline='') as file:
        writer = csv.writer(file, delimiter=';')

        writer.writerow([f'Missing Headers in {file_path_1}'])
        writer.writerows([[header] for header in missing_file1])

        writer.writerow([])

        writer.writerow([f'Missing Data Values in {file_path_1}'])
        writer.writerows([[value] for value in missing_file1_data])

    messagebox.showinfo("Success", f"Missing headers and missing data values have been saved to '{output_path}'.")


# Create tkinter window
root = tk.Tk()
root.title("File Comparison")
root.geometry("500x250")

# File 1
label_file1 = tk.Label(root, text="File 1:")
label_file1.grid(row=0, column=0, sticky="w")

entry_file1 = tk.Entry(root, width=50)
entry_file1.grid(row=0, column=1, padx=5)

button_browse_file1 = tk.Button(root, text="Browse", command=lambda: browse_file(entry_file1))
button_browse_file1.grid(row=0, column=2, padx=5)

# File 2
label_file2 = tk.Label(root, text="File 2:")
label_file2.grid(row=1, column=0, sticky="w")

entry_file2 = tk.Entry(root, width=50)
entry_file2.grid(row=1, column=1, padx=5)

button_browse_file2 = tk.Button(root, text="Browse", command=lambda: browse_file(entry_file2))
button_browse_file2.grid(row=1, column=2, padx=5)

# Delimiter1
label_delimiter1 = tk.Label(root, text="Delimiter File1:")
label_delimiter1.grid(row=2, column=0, sticky="w")

entry_delimiter1 = tk.Entry(root, width=5)
entry_delimiter1.insert(tk.END, ";")
entry_delimiter1.grid(row=2, column=1, padx=5)

# Delimiter2
label_delimiter2 = tk.Label(root, text="Delimiter File2:")
label_delimiter2.grid(row=3, column=0, sticky="w")

entry_delimiter2 = tk.Entry(root, width=5)
entry_delimiter2.insert(tk.END, ";")
entry_delimiter2.grid(row=3, column=1, padx=5)

# Header indices
label_header1 = tk.Label(root, text="Header Index (File 1):")
label_header1.grid(row=4, column=0, sticky="w")

entry_header1 = tk.Entry(root, width=5)
entry_header1.insert(tk.END, "0")
entry_header1.grid(row=4, column=1, padx=5)

label_header2 = tk.Label(root, text="Header Index (File 2):")
label_header2.grid(row=5, column=0, sticky="w")

entry_header2 = tk.Entry(root, width=5)
entry_header2.insert(tk.END, "0")
entry_header2.grid(row=5, column=1, padx=5)

# Output file
label_output_file = tk.Label(root, text="Output File:")
label_output_file.grid(row=6, column=0, sticky="w")

entry_output_file = tk.Entry(root, width=50)
entry_output_file.grid(row=6, column=1, padx=5)

button_browse_output_file = tk.Button(root, text="Browse", command=lambda: browse_output_file(entry_output_file))
button_browse_output_file.grid(row=6, column=2, padx=5)

# Process button
button_process = tk.Button(root, text="Process Files", command=process_files)
button_process.grid(row=7, column=0, columnspan=3, pady=10)

root.mainloop()
