import os
import requests
import json
import logging
import datetime
import sys
import openpyxl
import xml.etree.ElementTree as ET
import ast
import time
import uuid

TIMEOUT_DURATION = 30

Swag_sub_id = ''
Swag_feed_type = ''
Swag_dump_type = ''
Swag_perm_cont_id = ''
Swag_feed_id = ''
Swag_complexity = ''
Swag_recpt_name = ''
Swag_recpt_creation_failed = ''
Swag_recpt_status_code = ''
Swag_diss_creation_passed = ''
Swag_sub_proj_creation_passed = ''
Swag_fsq_filename_syntax = ''
Swag_fsq_filename_substr = ''
Swag_fsq_filename_complex = ''
Swag_encoding = ''
Swag_delimiter = ''
Swag_file_type = ''
Swag_delivery_type = ''
Swag_path = ''
Swag_host = ''
Swag_port = ''
Swag_username = ''
Swag_password = ''
Swag_recpt_count = ''
Swag_UnZip = ''
Swag_schedules = ''
Swag_hour = ''
Swag_minute = ''
Swag_time = ''
Swag_days = ''
Swag_alternative_email = ''
Swag_sender_email = ''
Swag_include_attachment = ''
Swag_project_guid = ''
feed_details_global = ''
Save_Status = ''
flag = 0
logging.basicConfig(filename='error.log', level=logging.ERROR, format='%(asctime)s:%(levelname)s:%(message)s')

user_path = r'C:\Users\Aditya.Apte\OneDrive - FE fundinfo\Desktop\Desktop Icons\Aditya Apte\FSQtoEFS'
EXCEL = ''
TRACKER = ''


def log_message(message):
    global user_path
    # Step 2: Write the message to the standard output
    print(str(message))
    # Step 3: Append the message to the log file
    # Step 2: Get the current date and time
    current_datetime = datetime.datetime.now()
    # Step 3: Format the date and time (optional)
    formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
    log_file_path = os.path.join(user_path, 'Output.log')
    with open(log_file_path, 'a') as log_file:
        log_file.write(formatted_datetime + ' - ' + str(message) + '\n')


def get_user():
    global user_path, EXCEL, TRACKER
    user_path = input("Please enter the OneDrive Path to 'FSQtoEFS'\nUser Path:")
    log_message("Thank you!")
    EXCEL = os.path.join(user_path, 'INDEX.xlsx')
    TRACKER = os.path.join(user_path, 'TRACKER.xlsx')
    log_message(f'Excel path built: {EXCEL}')
    log_message(f'Tracker path built: {TRACKER}')


# print("Opening file")
# wb = openpyxl.load_workbook(EXCEL)
# ws = wb['Sheet1']
#
# header1 = "Name"
# header2 = "Product"
# header3 = "Price"
#
# col_idx_name = get_column_index_by_header(header1)
# col_idx_product = get_column_index_by_header(header2)
# col_idx_price = get_column_index_by_header(header3)
#
# # Check if any header was not found
# if col_idx_price is None or col_idx_product is None or col_idx_name is None:
#     print("Could not find one or more headers")
# else:
#     print("Headers Found")
#     # Iterate through each row and check the condition for the 'Name' column values
#     for row in range(2, ws.max_row + 1):
#         cell_price = ws.cell(row=row, column=col_idx_price)
#         cell_product = ws.cell(row=row, column=col_idx_product)
#         print(cell_product.value)
#
#         # Example condition: Check if the value in 'Name' column is 'ConditionValue'
#         if cell_product.value == 'Carrots':
#             print(f"Row {row}: {cell_product.value} found, updating price.")
#             cell_price.value = 'Aditya'  # Update value based on condition
#         else:
#             print(f"Row {row}: {cell_product.value} does not match 'ConditionValue'")
#
#     # Save the workbook
#     wb.save(EXCEL)

def Script_Run():
    global flag, Swag_project_guid, Save_Status, feed_details_global, TRACKER, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time, flag, cell_minute, cell_hour

    run_count = 0
    counter = int(input("Please enter the number of iterations:"))
    if not 1 <= counter <= 1000:
        log_message("Number must be between 1 and 1000.")
        input("Hit ENTER to Exit!")
        sys.exit(1)
    log_message("Opening file")
    wb = openpyxl.load_workbook(filename=EXCEL, data_only=True)
    ws = wb['Bridge']

    wbt = openpyxl.load_workbook(filename=TRACKER, data_only=True)
    wst = wbt["Tracker"]
    max_row = wst.max_row
    while any(cell.value is not None for cell in wst[max_row]):
        max_row += 1

    track_row = max_row

    # Function to get the column index by header name
    def get_column_index_by_header(header):
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == header:
                return col[0].column
        return None

    def get_tracker_index_by_header(header):
        for col in wst.iter_cols(1, wst.max_column):
            if col[0].value == header:
                return col[0].column
        return None

    def time_to_minutes(time):
        hours, minutes = map(int, time.split(':'))
        return hours * 60 + minutes

    def get_EFS_bearer_token():
        log_message("Getting EFS Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'ssf-api-feed-read'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def get_Diss_bearer_token():
        log_message("Getting Diss Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'fefundinfo-esf-dissemination-api-dissemination-read'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def Search_Feed(input_feed_type, input_searchterm):
        global Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        log_message(f"Searching Feed: {input_searchterm}")
        url = f"https://datafeeds.fefundinfo.com/api/v1/Feeds/search"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }
        data = {
            "page": 1,
            "pageSize": 9,
            "orderBy": "feedStatus",
            "orderAscending": False,
            "filter": [
                {"key": "FeedDataType", "value": input_feed_type},
                {"key": "EngineType", "value": "SelfServedFeed"}
            ],
            "searchTerm": input_searchterm
        }

        response = requests.post(url, headers=headers, json=data)

        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            Swag_feed_id = data["payload"]["result"][0]["id"]
            Swag_perm_cont_id = data["payload"]["result"][0]["permissionContainerId"]

        else:
            log_message(f"SEARCH Request failed with status code {response.status_code}: {response.text}")

    def Trigger_FSQ_Diss(input_feed_id, input_sub_id):
        global Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        log_message("Running Swagger Script")
        url = f"https://fsqtoefsmigrationtooleuwliv.azurewebsites.net/FSQMigration/CreateOrUpdateDissemination?feedId={input_feed_id}&subscriptionId={input_sub_id}"
        headers = {
            'accept': '*/*'
        }


        response = requests.post(url, headers=headers)

        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            if "Recipient creation is failed" in data.get("payload", {}):
                log_message("Recipient Creation is Failed")
                Swag_recpt_creation_failed = "True"
                Swag_recpt_status_code = data["payload"]["Recipient status code"]
                Swag_diss_creation_passed = "False"
                Swag_sub_proj_creation_passed = "False"
            else:
                log_message("Recipient Creation Success")
                Swag_recpt_name = data["payload"]["Recipient Name"][0]
                Swag_recpt_creation_failed = "False"
                Swag_recpt_status_code = "Null"
                Swag_diss_creation_passed = data["payload"]["Dissemination Creation passed"]
                Swag_sub_proj_creation_passed = data["payload"]["Subscriptions project Creation passed"]

                return True
        else:
            log_message(f"Request failed with status code {response.status_code}: {response.text}")
            return False

    def Get_Feed_Details(input_feed_id):
        global feed_details_global, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        Swag_fsq_filename_complex = ''
        Swag_encoding = ''
        Swag_delimiter = ''
        Swag_file_type = ''
        Swag_delivery_type = ''
        Swag_path = ''
        Swag_host = ''
        Swag_port = ''
        Swag_username = ''
        Swag_password = ''
        Swag_UnZip = ''
        Swag_schedules = ''
        Swag_hour = ''
        Swag_minute = ''
        Swag_time = ''
        Swag_days = ''
        Swag_alternative_email = ''
        Swag_sender_email = ''
        Swag_include_attachment = ''
        log_message("Getting Feed Details")
        url = f"https://datafeeds.fefundinfo.com/api/v1/Feeds/{input_feed_id}"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
            'Cache-Control': 'no-cache',
            'Pragma': 'no-cache'
        }
        params = {'_': str(uuid.uuid4())}  # Generate a unique parameter to avoid caching
        response = requests.get(url, headers=headers, params=params)

        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            feed_details_global = data
            Swag_fsq_filename_complex = data["payload"]["feedFileSettings"].get("fileName", "")  # feed details
            Swag_encoding = data["payload"]["feedFileSettings"].get("encoding", "")  # feed details
            Swag_delimiter = data["payload"]["feedFileSettings"].get("separator", "")  # feed details
            Swag_file_type = data["payload"]["feedFileSettings"].get("fileType", "")  # feed details
            Swag_delivery_type = data["payload"]["deliverySettings"].get("method", "")  # feed details
            Swag_path = data["payload"]["deliverySettings"].get("path", "None")  # feed details
            Swag_host = data["payload"]["deliverySettings"].get("host", "None")  # feed details
            Swag_port = data["payload"]["deliverySettings"].get("port", "None")  # feed details
            Swag_username = data["payload"]["deliverySettings"].get("userName", "None")  # feed details
            Swag_password = data["payload"]["deliverySettings"].get("password", "None")  # feed details
            Swag_UnZip = data["payload"]["deliverySettings"].get("unZip", "")  # feed details
            if Swag_UnZip:
                Swag_UnZip = "YES"
            else:
                Swag_UnZip = "NO"
            Swag_schedules = data["payload"]["deliverySettings"].get("frequency", "")  # feed details
            Swag_hour = data["payload"]["deliverySettings"].get("hour", "")  # feed details
            Swag_minute = data["payload"]["deliverySettings"].get("minutes", "")  # feed details
            Swag_time = data["payload"]["deliverySettings"].get("time", "")  # feed details
            Swag_days = data["payload"]["deliverySettings"].get("day", "")  # feed details
            Swag_alternative_email = data["payload"]["deliverySettings"].get("emailAddress",
                                                                             "")  # feed details
            Swag_sender_email = data["payload"]["deliverySettings"].get("fromAddress", "")
            Swag_include_attachment = data["payload"]["deliverySettings"].get("attachFile", "")  # feed details
            if Swag_include_attachment:
                Swag_include_attachment = "YES"
            else:
                Swag_include_attachment = "NO"

        else:
            log_message(f"Request failed with status code {response.status_code}: {response.text}")

    def Save_Feed(payload):
        global Save_Status, feed_details_global, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        url = f"https://datafeeds.fefundinfo.com/api/v1/Feeds/save"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }
        data = payload

        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            Save_Status = data["payload"]["passed"]

        else:
            log_message(f"SAVE Request failed with status code {response.status_code}: {response.text}")

    def Get_Diss_Details(input_feed_id):
        global Swag_project_guid, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        log_message("Getting Dis Details")
        url = f"https://datafeeds.fefundinfo.com/api/v1/Feeds/{input_feed_id}/disseminationFeed"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        try:
            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                response_text = json.dumps(response.json(), indent=4)  # Format JSON output
                data = json.loads(response_text)
                Swag_project_guid = data["payload"][0].get("projectGuid", "")
            else:
                log_message(f"Request failed with status code {response.status_code}: {response.text}")
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            Swag_project_guid = 0

    def Get_recpt_count(proj_guid, perm_cont_id, diss_token):
        global Swag_project_guid, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        log_message("Getting Recpt count Details")
        url = f"https://efs-dissemination.fefundinfo.com/container/{perm_cont_id}/project/{proj_guid}/subscriber?Top=1000"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        try:
            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                response_text = json.dumps(response.json(), indent=4)  # Format JSON output
                data = json.loads(response_text)
                Swag_recpt_count = data.get('payload', {}).get('count', '')
            else:
                log_message(f"Request failed with status code {response.status_code}: {response.text}")
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            Swag_recpt_count = 0

    def Disable_Feed(input_feed_id):
        global Swag_project_guid, Swag_sub_id, Swag_feed_type, Swag_dump_type, Swag_perm_cont_id, Swag_feed_id, Swag_complexity, Swag_recpt_name, Swag_recpt_creation_failed, Swag_recpt_status_code, Swag_diss_creation_passed, Swag_sub_proj_creation_passed, Swag_fsq_filename_syntax, Swag_fsq_filename_substr, Swag_fsq_filename_complex, Swag_encoding, Swag_delimiter, Swag_file_type, Swag_delivery_type, Swag_path, Swag_host, Swag_port, Swag_username, Swag_password, Swag_recpt_count, Swag_UnZip, Swag_schedules, Swag_hour, Swag_minute, Swag_days, Swag_alternative_email, Swag_sender_email, Swag_include_attachment, Swag_time
        log_message("disabling feed")
        url = f"https://datafeeds.fefundinfo.com/api/v1/Feeds/{input_feed_id}/disable"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        try:
            response = requests.post(url, headers=headers)

            if response.status_code == 200:
                response_text = json.dumps(response.json(), indent=4)  # Format JSON output
                data = json.loads(response_text)
            else:
                log_message(f"Request failed with status code {response.status_code}: {response.text}")
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            Swag_project_guid = 0

    header_sub_id = "Sub_ID"
    header_dump_type = "Dump_Type"
    header_perm_cont_id = "Perm_Cont_ID"
    header_feed_id = "Feed_ID"
    header_complexity = "Complexity"
    header_recpt_name = "Recpt_Name"
    header_recpt_creation_failed = "Recpt_Creation_Failed"
    header_recpt_status_code = "Recpt_Status_Code"
    header_diss_creation_passed = "Diss_Creation_Passed"
    header_sub_proj_creation_passed = "Sub_Proj_Creation_Passed"
    header_fsq_filename_syntax = "FSQ_FileName_Syntax"
    header_fsq_filename_substr = "FSQ_FileName_SubStr"
    header_fsq_filename_complex = "FSQ_FileName_Complex"
    header_encoding = "Encoding"
    header_delimiter = "Delimiter"
    header_file_type = "File_Type"
    header_delivery_type = "Delivery_Type"
    header_path = "Path"
    header_host = "Host"
    header_port = "Port"
    header_username = "Username"
    header_password = "Password"
    header_recpt_count = "Recpt_Count"
    header_compressed = "UnZip"
    header_schedules = "Schedules"
    header_time = "Time"
    header_alternative_email = "Alternative_Email"
    header_sender_email = "Sender_Email"
    header_include_attachment = "Include_Attachment"
    header_status = "Status"

    col_idx_sub_id = get_column_index_by_header(header_sub_id)
    col_idx_dump_type = get_column_index_by_header(header_dump_type)
    col_idx_perm_cont_id = get_column_index_by_header(header_perm_cont_id)
    col_idx_feed_id = get_column_index_by_header(header_feed_id)
    col_idx_complexity = get_column_index_by_header(header_complexity)
    col_idx_recpt_name = get_column_index_by_header(header_recpt_name)
    col_idx_recpt_creation_failed = get_column_index_by_header(header_recpt_creation_failed)
    col_idx_recpt_status_code = get_column_index_by_header(header_recpt_status_code)
    col_idx_diss_creation_passed = get_column_index_by_header(header_diss_creation_passed)
    col_idx_sub_proj_creation_passed = get_column_index_by_header(header_sub_proj_creation_passed)
    col_idx_fsq_filename_syntax = get_column_index_by_header(header_fsq_filename_syntax)
    col_idx_fsq_filename_substr = get_column_index_by_header(header_fsq_filename_substr)
    col_idx_fsq_filename_complex = get_column_index_by_header(header_fsq_filename_complex)
    col_idx_encoding = get_column_index_by_header(header_encoding)
    col_idx_delimiter = get_column_index_by_header(header_delimiter)
    col_idx_file_type = get_column_index_by_header(header_file_type)
    col_idx_delivery_type = get_column_index_by_header(header_delivery_type)
    col_idx_path = get_column_index_by_header(header_path)
    col_idx_host = get_column_index_by_header(header_host)
    col_idx_port = get_column_index_by_header(header_port)
    col_idx_username = get_column_index_by_header(header_username)
    col_idx_password = get_column_index_by_header(header_password)
    col_idx_recpt_count = get_column_index_by_header(header_recpt_count)
    col_idx_compressed = get_column_index_by_header(header_compressed)
    col_idx_schedules = get_column_index_by_header(header_schedules)
    col_idx_time = get_column_index_by_header(header_time)
    col_idx_alternative_email = get_column_index_by_header(header_alternative_email)
    col_idx_sender_email = get_column_index_by_header(header_sender_email)
    col_idx_include_attachment = get_column_index_by_header(header_include_attachment)
    col_idx_status = get_column_index_by_header(header_status)

    tracker_sub_id = "Sub_ID"
    tracker_feed_type = "Feed_Type"
    tracker_dump_type = "Dump_Type"
    tracker_perm_cont_id = "Perm_Cont_ID"
    tracker_feed_id = "Feed_ID"
    tracker_complexity = "Complexity"
    tracker_recpt_name = "Recpt_Name"
    tracker_recpt_creation_failed = "Recpt_Creation_Failed"
    tracker_recpt_status_code = "Recpt_Status_Code"
    tracker_diss_creation_passed = "Diss_Creation_Passed"
    tracker_sub_proj_creation_passed = "Sub_Proj_Creation_Passed"
    tracker_fsq_filename_complex_match = "FSQ_FileName_Complex_Match"
    tracker_encoding_match = "Encoding_Match"
    tracker_delimiter_match = "Delimiter_Match"
    tracker_file_type_match = "File_Type_Match"
    tracker_delivery_type_match = "Delivery_Type_Match"
    tracker_path_match = "Path_Match"
    tracker_host_match = "Host_Match"
    tracker_port_match = "Port_Match"
    tracker_username_match = "Username_Match"
    tracker_password_match = "Password_Match"
    tracker_recpt_count_match = "Recpt_Count_Match"
    tracker_compressed_match = "Compressed_Match"
    tracker_schedules_match = "Schedules_Match"
    tracker_days_match = "Days_Match"
    tracker_time_match = "Time_Match"
    tracker_alternative_email_match = "Alternative_Email_Match"
    tracker_sender_email_match = "Sender_Email_Match"
    tracker_include_attachment_match = "Include_Attachment_Match"
    tracker_status = "Status"

    tracker_idx_sub_id = get_tracker_index_by_header(tracker_sub_id)
    tracker_idx_feed_type = get_tracker_index_by_header(tracker_feed_type)
    tracker_idx_dump_type = get_tracker_index_by_header(tracker_dump_type)
    tracker_idx_perm_cont_id = get_tracker_index_by_header(tracker_perm_cont_id)
    tracker_idx_feed_id = get_tracker_index_by_header(tracker_feed_id)
    tracker_idx_complexity = get_tracker_index_by_header(tracker_complexity)
    tracker_idx_recpt_name = get_tracker_index_by_header(tracker_recpt_name)
    tracker_idx_recpt_creation_failed = get_tracker_index_by_header(tracker_recpt_creation_failed)
    tracker_idx_recpt_status_code = get_tracker_index_by_header(tracker_recpt_status_code)
    tracker_idx_diss_creation_passed = get_tracker_index_by_header(tracker_diss_creation_passed)
    tracker_idx_sub_proj_creation_passed = get_tracker_index_by_header(tracker_sub_proj_creation_passed)
    tracker_idx_fsq_filename_complex_match = get_tracker_index_by_header(tracker_fsq_filename_complex_match)
    tracker_idx_encoding_match = get_tracker_index_by_header(tracker_encoding_match)
    tracker_idx_delimiter_match = get_tracker_index_by_header(tracker_delimiter_match)
    tracker_idx_file_type_match = get_tracker_index_by_header(tracker_file_type_match)
    tracker_idx_delivery_type_match = get_tracker_index_by_header(tracker_delivery_type_match)
    tracker_idx_path_match = get_tracker_index_by_header(tracker_path_match)
    tracker_idx_host_match = get_tracker_index_by_header(tracker_host_match)
    tracker_idx_port_match = get_tracker_index_by_header(tracker_port_match)
    tracker_idx_username_match = get_tracker_index_by_header(tracker_username_match)
    tracker_idx_password_match = get_tracker_index_by_header(tracker_password_match)
    tracker_idx_recpt_count_match = get_tracker_index_by_header(tracker_recpt_count_match)
    tracker_idx_compressed_match = get_tracker_index_by_header(tracker_compressed_match)
    tracker_idx_schedules_match = get_tracker_index_by_header(tracker_schedules_match)
    tracker_idx_days_match = get_tracker_index_by_header(tracker_days_match)
    tracker_idx_time_match = get_tracker_index_by_header(tracker_time_match)
    tracker_idx_alternative_email_match = get_tracker_index_by_header(tracker_alternative_email_match)
    tracker_idx_sender_email_match = get_tracker_index_by_header(tracker_sender_email_match)
    tracker_idx_include_attachment_match = get_tracker_index_by_header(tracker_include_attachment_match)
    tracker_idx_status = get_tracker_index_by_header(tracker_status)

    token = get_EFS_bearer_token()
    diss_token = get_Diss_bearer_token()
    print("=====================================================")
    print("=====================================================")

    for row in range(2, ws.max_row + 1):
        cell_status = ws.cell(row=row, column=col_idx_status)
        if cell_status.value is None or cell_status.value == '':
            if run_count == counter:
                break
            run_count += 1
            print("=====================================================")
            print("=====================================================")
            log_message(f'Run Count= {run_count}')
            cell_sub_id = ws.cell(row=row, column=col_idx_sub_id).value
            cell_dump_type = ws.cell(row=row, column=col_idx_dump_type).value
            log_message(f'Cell Dump type: {cell_dump_type}')
            if cell_dump_type == "Documentary data":
                cell_feed_type = "Document"
            elif cell_dump_type == "Standard Static data" or cell_dump_type == "Dynamic data":
                cell_feed_type = "FundData"
            else:
                log_message("Feed Type not Found!")
                input("Hit Enter to Exit!")
                exit(0)

            cell_complexity = ws.cell(row=row, column=col_idx_complexity).value
            cell_fsq_filename_syntax = ws.cell(row=row, column=col_idx_fsq_filename_syntax).value
            cell_fsq_filename_substr = ws.cell(row=row, column=col_idx_fsq_filename_substr).value
            if cell_fsq_filename_substr in [None, 0]:
                cell_fsq_filename_substr = ''
            cell_fsq_filename_complex = ws.cell(row=row, column=col_idx_fsq_filename_complex).value
            cell_encoding = ws.cell(row=row, column=col_idx_encoding).value
            cell_delimiter = ws.cell(row=row, column=col_idx_delimiter).value
            cell_file_type = ws.cell(row=row, column=col_idx_file_type).value
            cell_file_type = cell_file_type.strip()
            cell_delivery_type = ws.cell(row=row, column=col_idx_delivery_type).value
            cell_path = ws.cell(row=row, column=col_idx_path).value or "None"
            cell_host = ws.cell(row=row, column=col_idx_host).value or "None"
            cell_port = ws.cell(row=row, column=col_idx_port).value or "None"
            cell_username = ws.cell(row=row, column=col_idx_username).value or "None"
            cell_password = ws.cell(row=row, column=col_idx_password).value or "None"
            cell_recpt_count = ws.cell(row=row, column=col_idx_recpt_count).value
            cell_UnZip = ws.cell(row=row, column=col_idx_compressed).value
            cell_schedules = ws.cell(row=row, column=col_idx_schedules).value
            cell_days = ''
            cell_time = ws.cell(row=row, column=col_idx_time).value
            cell_alternative_email = ws.cell(row=row, column=col_idx_alternative_email).value
            cell_sender_email = ws.cell(row=row, column=col_idx_sender_email).value
            cell_include_attachment = ws.cell(row=row, column=col_idx_include_attachment).value
            cell_status = ws.cell(row=row, column=col_idx_status)
            cell_hour = 0
            cell_minute = 0

            Tracker_sub_id_value = wst.cell(row=track_row, column=tracker_idx_sub_id)
            Tracker_feed_type_value = wst.cell(row=track_row, column=tracker_idx_feed_type)
            Tracker_dump_type_value = wst.cell(row=track_row, column=tracker_idx_dump_type)
            Tracker_perm_cont_id_value = wst.cell(row=track_row, column=tracker_idx_perm_cont_id)
            Tracker_feed_id_value = wst.cell(row=track_row, column=tracker_idx_feed_id)
            Tracker_complexity_value = wst.cell(row=track_row, column=tracker_idx_complexity)
            Tracker_recpt_name_value = wst.cell(row=track_row, column=tracker_idx_recpt_name)
            Tracker_recpt_creation_failed_value = wst.cell(row=track_row, column=tracker_idx_recpt_creation_failed)
            Tracker_recpt_status_code_value = wst.cell(row=track_row, column=tracker_idx_recpt_status_code)
            Tracker_diss_creation_passed_value = wst.cell(row=track_row, column=tracker_idx_diss_creation_passed)
            Tracker_sub_proj_creation_passed_value = wst.cell(row=track_row,
                                                              column=tracker_idx_sub_proj_creation_passed)
            Tracker_fsq_filename_complex_match_value = wst.cell(row=track_row,
                                                                column=tracker_idx_fsq_filename_complex_match)
            Tracker_encoding_match_value = wst.cell(row=track_row, column=tracker_idx_encoding_match)
            Tracker_delimiter_match_value = wst.cell(row=track_row, column=tracker_idx_delimiter_match)
            Tracker_file_type_match_value = wst.cell(row=track_row, column=tracker_idx_file_type_match)
            Tracker_delivery_type_match_value = wst.cell(row=track_row, column=tracker_idx_delivery_type_match)
            Tracker_path_match_value = wst.cell(row=track_row, column=tracker_idx_path_match)
            Tracker_host_match_value = wst.cell(row=track_row, column=tracker_idx_host_match)
            Tracker_port_match_value = wst.cell(row=track_row, column=tracker_idx_port_match)
            Tracker_username_match_value = wst.cell(row=track_row, column=tracker_idx_username_match)
            Tracker_password_match_value = wst.cell(row=track_row, column=tracker_idx_password_match)
            Tracker_recpt_count_match_value = wst.cell(row=track_row, column=tracker_idx_recpt_count_match)
            Tracker_compressed_match_value = wst.cell(row=track_row, column=tracker_idx_compressed_match)
            Tracker_schedules_match_value = wst.cell(row=track_row, column=tracker_idx_schedules_match)
            Tracker_days_match_value = wst.cell(row=track_row, column=tracker_idx_days_match)
            Tracker_time_match_value = wst.cell(row=track_row, column=tracker_idx_time_match)
            Tracker_alternative_email_match_value = wst.cell(row=track_row, column=tracker_idx_alternative_email_match)
            Tracker_sender_email_match_value = wst.cell(row=track_row, column=tracker_idx_sender_email_match)
            Tracker_include_attachment_match_value = wst.cell(row=track_row,
                                                              column=tracker_idx_include_attachment_match)
            Tracker_status_value = wst.cell(row=track_row, column=tracker_idx_status)

            if cell_fsq_filename_syntax.startswith('<?xml'):
                root = ET.fromstring(cell_fsq_filename_syntax)
                pattern = root.find('.//Pattern').text
                cell_fsq_filename_complex = pattern
                if 'CdEchantAbont' in pattern:
                    cell_fsq_filename_complex = pattern.replace('{CdEchantAbont}', cell_fsq_filename_substr)
                if '{IdDiffus}.{SuffixFich}' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('{IdDiffus}.{SuffixFich}', '{feedid}')
                if '{IdDiffus}' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('{IdDiffus}', '{feedid}')
                if '{SuffixFich}' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('.{SuffixFich}', '')
                if '{DtHr}' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('{DtHr}', '{yyyMMddHHmmss}')

            if cell_fsq_filename_syntax.startswith('DEFAULT:'):
                cell_fsq_filename_complex = cell_fsq_filename_syntax.replace('DEFAULT: ', '').replace('CdEchantAbont',
                                                                                                      cell_fsq_filename_substr)
                cell_fsq_filename_complex = cell_fsq_filename_complex.replace('-yyyyMMddHHmmss-', '-{yyyyMMddHHmmss}-')
                if 'IdDiffus.SuffixFich' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('IdDiffus.SuffixFich', '{feedid}')
                if 'IdDiffus' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('IdDiffus', '{feedid}')
                if 'SuffixFich' in cell_fsq_filename_complex:
                    cell_fsq_filename_complex = cell_fsq_filename_complex.replace('SuffixFich', '{feedid}')

            # Encoding Mapping Here
            encoding_map = {
                "UTF-8": 1,
                "ISO8859_1": 5,
                "UTF8": 0,
                "ASCII": 2,
                "Cp1252": 4,
                "ISO8859_2": 6
            }

            swag_encoding = encoding_map.get(cell_encoding)

            if swag_encoding is not None:
                cell_encoding = swag_encoding

            # Delimiter Mapping Here
            delimiter_map = {
                "Semi-Colon (;)": ";",
                "Comma (,)": ",",
                "Pipe (|)": "|"
            }
            swag_delimiter = delimiter_map.get(cell_delimiter)

            if swag_delimiter is not None:
                cell_delimiter = swag_delimiter

            # Delivery Method Mapping Here
            if cell_delivery_type == "MZIP":
                cell_delivery_type = 3
                cell_UnZip = "NO"
            else:
                delivery_type_map = {
                    "FTP": 2,
                    "HTTP": 4,
                    "MAIL": 3,
                    "SFTP": 1
                }
                swag_delivery_type = delivery_type_map.get(cell_delivery_type)

                if swag_delivery_type is not None:
                    cell_delivery_type = swag_delivery_type

            # Days Mapping Here
            days_map = {
                "Daily (week-ends excluded)": "['1', '2', '3', '4', '5']",
                "Weekly days: Mon,Tue,Wed,Thu,Fri": "['1', '2', '3', '4', '5']",
                "Monthly : 21": "[21]",
                "Monthly : 2": "[2]",
                "Weekly days: Mon": "[1]",
                "Monthly : 5": "[5]",
                "Monthly : 15": "[15]",
                "Weekly days: Tue,Wed,Thu,Fri": "['2', '3', '4', '5']",
                "Weekly days: Mon,Tue,Wed,Thu,Fri,Sat": "['1', '2', '3', '4', '5', '6']",
                "Weekly days: Mon,Tue,Wed,Thu": "['1', '2', '3', '4']",
                "Monthly : 1": "[1]",
                "Monthly : 3, 7": "['3', '7']",
                "Weekly days: Wed": "[3]",
                "Weekly days: Thu": "[4]",
                "Weekly days: Sat": "[6]"
            }
            swag_days = days_map.get(cell_schedules)

            if swag_days is not None:
                cell_days = swag_days
                if len(cell_days) > 4:
                    cell_days = ast.literal_eval(cell_days)

            # Schedules Mapping Here
            schedules_map = {
                "Daily (week-ends excluded)": 2,
                "Weekly days: Mon,Tue,Wed,Thu,Fri": 2,
                "Monthly : 21": 3,
                "Monthly : 2": 3,
                "Weekly days: Mon": 2,
                "Monthly : 5": 3,
                "Monthly : 15": 3,
                "Weekly days: Tue,Wed,Thu,Fri": 2,
                "Weekly days: Mon,Tue,Wed,Thu,Fri,Sat": 2,
                "Weekly days: Mon,Tue,Wed,Thu": 2,
                "Monthly : 1": 3,
                "Monthly : 3, 7": 3,
                "Weekly days: Wed": 2,
                "Weekly days: Thu": 2,
                "Weekly days: Sat": 2,
                "Daily (week-ends included)": 1
            }
            swag_schedules = schedules_map.get(cell_schedules)

            if swag_schedules is not None:
                cell_schedules = swag_schedules
            try:
                if len(cell_time) > 8:
                    flag = 0
                    cell_schedules = 0
                    cell_time = cell_time.split(", ")
                    cell_time = sorted(time_to_minutes(t) for t in cell_time)
                    cell_time = [str(minute) for minute in cell_time]
            except Exception as e:
                cell_time = cell_time.strftime("%H:%M:%S")
                cell_hour, cell_minute, _ = cell_time.split(':')
                cell_hour = int(cell_hour)
                cell_minute = int(cell_minute)
                if cell_minute % 5 == 0 and cell_minute % 10 != 0:
                    cell_minute = cell_minute - (cell_minute % 10)
                flag = 1

            Swag_fsq_filename_complex = ''  # feed details
            Swag_perm_cont_id = ''  # feed search
            Swag_feed_id = ''  # feed search
            Swag_complexity = ''
            Swag_recpt_name = ''  # feed diss trigger
            Swag_recpt_creation_failed = ''  # feed diss trigger
            Swag_recpt_status_code = ''  # feed diss trigger
            Swag_diss_creation_passed = ''  # feed diss trigger
            Swag_sub_proj_creation_passed = ''  # feed diss trigger
            Swag_encoding = ''  # feed details
            Swag_delimiter = ''  # feed details
            Swag_file_type = ''  # feed details
            Swag_delivery_type = ''  # feed details
            Swag_path = ''  # feed details
            Swag_host = ''  # feed details
            Swag_port = ''  # feed details
            Swag_username = ''  # feed details
            Swag_password = ''  # feed details
            Swag_recpt_count = ''  # feed diss details
            Swag_UnZip = ''  # feed details
            Swag_schedules = ''  # feed details
            Swag_hour = ''  # feed details
            Swag_alternative_email = ''  # feed details
            Swag_sender_email = ''  # feed details
            Swag_include_attachment = ''  # feed details

            # MAJOR RUNS HAPPEN HERE
            search_term_build = f'FSQ-{cell_sub_id}'
            log_message(f"Running Sub ID: {cell_sub_id}")
            Search_Feed(cell_feed_type, search_term_build)
            Run_Status = Trigger_FSQ_Diss(Swag_feed_id, cell_sub_id)
            log_message(f'Run Status: {Run_Status}')
            Get_Feed_Details(Swag_feed_id)
            Get_Diss_Details(Swag_feed_id)
            Get_recpt_count(Swag_project_guid, Swag_perm_cont_id, diss_token)
            if Swag_delivery_type in (2, 1):
                Disable_Feed(Swag_feed_id)
            # API CALLS END HERE

            if Swag_schedules != 0:
                Swag_time = f'{Swag_hour:02}:{Swag_minute:02}:00'
            print("=====================================================")
            print("=====================================================")
            print("STARTING COMPARISON")
            Tracker_sub_id_value.value = cell_sub_id
            Tracker_feed_type_value.value = cell_feed_type
            Tracker_dump_type_value.value = cell_dump_type
            Tracker_perm_cont_id_value.value = Swag_perm_cont_id
            Tracker_feed_id_value.value = Swag_feed_id
            Tracker_complexity_value.value = cell_complexity
            Tracker_recpt_name_value.value = Swag_recpt_name
            Tracker_recpt_creation_failed_value.value = Swag_recpt_creation_failed
            Tracker_recpt_status_code_value.value = Swag_recpt_status_code
            Tracker_diss_creation_passed_value.value = Swag_diss_creation_passed
            Tracker_sub_proj_creation_passed_value.value = Swag_sub_proj_creation_passed

            trackers1 = [
                ('filename', cell_fsq_filename_complex, Swag_fsq_filename_complex,
                 Tracker_fsq_filename_complex_match_value),
                ('encoding', cell_encoding, Swag_encoding, Tracker_encoding_match_value),
                ('delimiter', cell_delimiter, Swag_delimiter, Tracker_delimiter_match_value),
                ('file_type', cell_file_type, Swag_file_type, Tracker_file_type_match_value),
                ('delivery_type', cell_delivery_type, Swag_delivery_type, Tracker_delivery_type_match_value),
                ('path', cell_path, Swag_path, Tracker_path_match_value),
                ('host', cell_host, Swag_host, Tracker_host_match_value),
                ('port', cell_port, Swag_port, Tracker_port_match_value),
                ('username', cell_username, Swag_username, Tracker_username_match_value),
                ('password', cell_password, Swag_password, Tracker_password_match_value),
                ('recpt_count', cell_recpt_count, Swag_recpt_count, Tracker_recpt_count_match_value),
                ('compressed', cell_UnZip, Swag_UnZip, Tracker_compressed_match_value),
                ('schedules', cell_schedules, Swag_schedules, Tracker_schedules_match_value),
                ('days', cell_days, Swag_days, Tracker_days_match_value),
                ('time', cell_time, Swag_time, Tracker_time_match_value),
                ('alternative_email', cell_alternative_email, Swag_alternative_email,
                 Tracker_alternative_email_match_value),
                ('sender_email', cell_sender_email, Swag_sender_email, Tracker_sender_email_match_value),
                ('include_attachment', cell_include_attachment, Swag_include_attachment,
                 Tracker_include_attachment_match_value)
            ]

            trackers2 = [
                ('filename', cell_fsq_filename_complex, Swag_fsq_filename_complex,
                 Tracker_fsq_filename_complex_match_value),
                ('encoding', cell_encoding, Swag_encoding, Tracker_encoding_match_value),
                ('delimiter', cell_delimiter, Swag_delimiter, Tracker_delimiter_match_value),
                ('file_type', cell_file_type, Swag_file_type, Tracker_file_type_match_value),
                ('delivery_type', cell_delivery_type, Swag_delivery_type, Tracker_delivery_type_match_value),
                ('path', cell_path, Swag_path, Tracker_path_match_value),
                ('host', cell_host, Swag_host, Tracker_host_match_value),
                ('port', cell_port, Swag_port, Tracker_port_match_value),
                ('username', cell_username, Swag_username, Tracker_username_match_value),
                ('password', cell_password, Swag_password, Tracker_password_match_value),
                ('recpt_count', cell_recpt_count, Swag_recpt_count, Tracker_recpt_count_match_value),
                ('compressed', cell_UnZip, Swag_UnZip, Tracker_compressed_match_value),
                ('schedules', cell_schedules, Swag_schedules, Tracker_schedules_match_value),
                ('days', cell_days, Swag_days, Tracker_days_match_value),
                ('hour', cell_hour, Swag_hour, Tracker_time_match_value),
                ('minute', cell_minute, Swag_minute, Tracker_time_match_value),
                ('alternative_email', cell_alternative_email, Swag_alternative_email,
                 Tracker_alternative_email_match_value),
                ('sender_email', cell_sender_email, Swag_sender_email, Tracker_sender_email_match_value),
                ('include_attachment', cell_include_attachment, Swag_include_attachment,
                 Tracker_include_attachment_match_value)
            ]

            # Data field mappings
            data_mappings1 = {
                'filename': 'payload.feedFileSettings.fileName',
                'encoding': 'payload.feedFileSettings.encoding',
                'delimiter': 'payload.feedFileSettings.separator',
                'file_type': 'payload.feedFileSettings.fileType',
                'delivery_type': 'payload.deliverySettings.method',
                'path': 'payload.deliverySettings.path',
                'host': 'payload.deliverySettings.host',
                'port': 'payload.deliverySettings.port',
                'username': 'payload.deliverySettings.userName',
                'password': 'payload.deliverySettings.password',
                'recpt_count': 'payload.uploadedIdentifiersSuccessCount',
                'compressed': 'payload.deliverySettings.unZip',
                'schedules': 'payload.deliverySettings.frequency',
                'days': 'payload.deliverySettings.day',
                'time': 'payload.deliverySettings.time',
                'alternative_email': 'payload.deliverySettings.useAlternateEmailAddress',
                'sender_email': 'payload.deliverySettings.fromAddress',
                'include_attachment': 'payload.deliverySettings.attachFile'
            }

            data_mappings2 = {
                'filename': 'payload.feedFileSettings.fileName',
                'encoding': 'payload.feedFileSettings.encoding',
                'delimiter': 'payload.feedFileSettings.separator',
                'file_type': 'payload.feedFileSettings.fileType',
                'delivery_type': 'payload.deliverySettings.method',
                'path': 'payload.deliverySettings.path',
                'host': 'payload.deliverySettings.host',
                'port': 'payload.deliverySettings.port',
                'username': 'payload.deliverySettings.userName',
                'password': 'payload.deliverySettings.password',
                'recpt_count': 'payload.uploadedIdentifiersSuccessCount',
                'compressed': 'payload.deliverySettings.unZip',
                'schedules': 'payload.deliverySettings.frequency',
                'days': 'payload.deliverySettings.day',
                'hour': 'payload.deliverySettings.hour',
                'minute': 'payload.deliverySettings.minutes',
                'alternative_email': 'payload.deliverySettings.useAlternateEmailAddress',
                'sender_email': 'payload.deliverySettings.fromAddress',
                'include_attachment': 'payload.deliverySettings.attachFile'
            }

            # Function to set the value in nested dictionary
            def set_nested_value(data, keys, value):
                keys = keys.split('.')
                for key in keys[:-1]:
                    data = data[key]
                data[keys[-1]] = value

            # Function to convert YES/NO to True/False
            def convert_to_bool(value):
                return value == 'YES'

            # Update values where condition is false
            all_true = True
            if flag == 0:
                for heading, cell_value, swag_value, tracker in trackers1:
                    if cell_value != swag_value:
                        tracker.value = "False"
                        if heading == "file_type" and (cell_value == "txt" or cell_value == "pdf"):
                            tracker.value = "True"
                        else:
                            key_chain = data_mappings1[heading]
                            if heading in ['compressed', 'include_attachment']:
                                cell_value = convert_to_bool(cell_value)
                            set_nested_value(feed_details_global, key_chain, cell_value)
                            all_true = False
                    else:
                        tracker.value = "True"
                    # print(f"{heading}:\n{cell_value}\n{swag_value}\n{tracker.value}\n================")
                    # time.sleep(0.25)
            elif flag == 1:
                print("Exe Flag 1")
                for heading, cell_value, swag_value, tracker in trackers2:
                    if cell_value != swag_value:
                        tracker.value = "False"
                        if heading == "file_type" and (cell_value == "txt" or cell_value == "pdf"):
                            tracker.value = "True"
                        else:
                            key_chain = data_mappings2[heading]
                            if heading in ['compressed', 'include_attachment']:
                                cell_value = convert_to_bool(cell_value)
                            set_nested_value(feed_details_global, key_chain, cell_value)
                            all_true = False
                    else:
                        tracker.value = "True"
                    # print(f"{heading}:\n{cell_value}\n{swag_value}\n{tracker.value}\n================")
                    # time.sleep(0.25)

            print("COMPARISON COMPLETED")
            print(f'All True: {all_true}')

            print("=====================================================")
            print("=====================================================")
            feed_details_global = feed_details_global["payload"]
            if not all_true:
                print("Rewriting Feed")
                feed_details_global["deliverySettings"]["emailNotification"] = True
                feed_details_global["deliverySettings"]["useAlternateEmailAddress"] = True
                feed_details_global["deliverySettings"]["alternateEmailConsent"] = True
                feed_details_global["deliverySettings"]["emailAddress"] = "efsdelivery@fefundinfo.com"
                Save_Feed(feed_details_global)
                #time.sleep(2)
                log_message("Saving Feed - Please wait")
                #time.sleep(2)
                print("Validating Save ...")
                #time.sleep(2)
                print(f'Save Status Response API : {Save_Status}')
                #time.sleep(2)
                # HERE
                track_row += 1
                Tracker_status_value.value = "Error"
                wbt.save(TRACKER)
                wb.save(EXCEL)
            else:
                cell_status.value = "Success"
                Tracker_status_value.value = "Success"
                track_row += 1
                wbt.save(TRACKER)
                wb.save(EXCEL)
                print("ALL TRUE")

                # HERE
            #     Get_Feed_Details(Swag_feed_id)
            #     all_true = True
            #     if flag == 0:
            #         for heading, cell_value, swag_value, tracker in trackers1:
            #             if cell_value != swag_value:
            #                 tracker.value = "False"
            #                 if heading == "file_type" and cell_value == "txt":
            #                     tracker.value = "True"
            #                 else:
            #                     key_chain = data_mappings1[heading]
            #                     if heading in ['compressed', 'include_attachment']:
            #                         cell_value = convert_to_bool(cell_value)
            #                     set_nested_value(feed_details_global, key_chain, cell_value)
            #                     all_true = False
            #             else:
            #                 tracker.value = "True"
            #             print(f"{heading}:\n{cell_value}\n{swag_value}\n{tracker.value}\n================")
            #     elif flag == 1:
            #         for heading, cell_value, swag_value, tracker in trackers2:
            #             if cell_value != swag_value:
            #                 tracker.value = "False"
            #                 if heading == "file_type" and cell_value == "txt":
            #                     tracker.value = "True"
            #                 else:
            #                     key_chain = data_mappings2[heading]
            #                     if heading in ['compressed', 'include_attachment']:
            #                         cell_value = convert_to_bool(cell_value)
            #                     set_nested_value(feed_details_global, key_chain, cell_value)
            #                     all_true = False
            #             else:
            #                 tracker.value = "True"
            #             print(f"{heading}:\n{cell_value}\n{swag_value}\n{tracker.value}\n================")
            #     if all_true:
            #         print("ALL TRUE")
            #         input("Hit Enter to continue run!")
            #     else:
            #         print("FALSE")
            # else:
            #     cell_status.value = "Success"
            #     Tracker_status_value.value = "Success"
            #     track_row += 1
            #     wbt.save(TRACKER)
            #     wb.save(EXCEL)
            #     print("ALL TRUE")
            # if not all_true:
            #     log_message(f'Sub ID: {cell_sub_id} did not pass!')
            #     wbt.save(TRACKER)
            #     wb.save(EXCEL)
            #     input("Hit Enter to Exit!")
            #     exit(0)


get_user()
Script_Run()
