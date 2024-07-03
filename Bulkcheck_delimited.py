import os
import pandas as pd
import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import messagebox
from tqdm import tqdm


def read_csv_file(filename, delimiter=';'):
    rows = []
    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            rows.append(row)
    return rows


def read_xlsx_file(filename):
    rows = []
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows


def find_missing_headers(file1_headers, file2_headers):
    missing_file1 = set(file2_headers) - set(file1_headers)
    missing_file2 = set(file1_headers) - set(file2_headers)
    return missing_file1, missing_file2


def find_column_index(header_row, column_name):
    return header_row.index(column_name) if column_name in header_row else -1


def find_missing_data(file1_data, file2_data, column_index1, column_index2):
    file1_values = set(row[column_index1] for row in tqdm(file1_data[1:], desc='Processing File 1', unit='row',
                                                          ncols=80)) if column_index1 >= 0 else set()
    file2_values = set(row[column_index2] for row in tqdm(file2_data[1:], desc='Processing File 2', unit='row',
                                                          ncols=80)) if column_index2 >= 0 else set()
    missing_file1 = file2_values - file1_values
    missing_file2 = file1_values - file2_values
    return missing_file1, missing_file2


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def browse_output_file(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[('CSV Files', '*.csv')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def process_files():
    # Get file path
    file_path = entry_file1.get()

    # Check if file is selected
    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return

    # Read Excel file
    try:
        df = pd.read_excel(file_path)
    except:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return

    # Get the directory path values
    try:
        paths = df["Path"].tolist()
        delimiters = df["Delimiter"].tolist()
    except KeyError:
        messagebox.showerror("Error", "The 'path' column does not exist in the Excel file.")
        return

    # Iterate over the paths
    for path, delimiter in zip(paths, delimiters):
        # Check if the path is valid
        if os.path.exists(path):
            # Get the list of files in the directory
            files = os.listdir(path)
            # Assuming the first two files in the directory are file 1 and file 2
            if len(files) >= 2:
                file_pathx = os.path.join(path, files[0])
                file_pathy = os.path.join(path, files[1])
                if file_pathx.lower().endswith('.csv') and file_pathy.lower().endswith('.xlsx'):
                    file1_path = os.path.join(path, files[0])
                    file1_data = read_csv_file(file1_path, delimiter)
                    file2_path = os.path.join(path, files[1])
                    file2_data = read_xlsx_file(file2_path)
                elif file_pathx.lower().endswith('.xlsx') and file_pathy.lower().endswith('.csv'):
                    file1_path = os.path.join(path, files[1])
                    file1_data = read_csv_file(file1_path, delimiter)
                    file2_path = os.path.join(path, files[0])
                    file2_data = read_xlsx_file(file2_path)
                elif file_pathx.lower().endswith('.csv') and file_pathy.lower().endswith('.csv'):
                    file1_path = os.path.join(path, files[1])
                    file1_data = read_csv_file(file1_path, delimiter)
                    file2_path = os.path.join(path, files[0])
                    file2_data = read_csv_file(file2_path, delimiter)
                else:
                    print(path)
                    continue

                # Check if headers are present
                if not file1_data or not file2_data:
                    messagebox.showwarning("Warning", "One or both files are empty.")
                    continue

                file1_headers = file1_data[0]
                file2_headers = file2_data[0]

                # Find missing headers
                missing_file1_headers, missing_file2_headers = find_missing_headers(file1_headers, file2_headers)

                # Find missing data
                column_index1 = find_column_index(file1_headers, "OFST020000")
                column_index2 = find_column_index(file2_headers, "OFST020000")
                missing_file1_data, missing_file2_data = find_missing_data(file1_data, file2_data, column_index1,
                                                                           column_index2)

                # Write the results to an output file

                output_path = entry_output.get()
                with open(output_path, 'a', encoding='utf-8', newline='') as file:
                    writer = csv.writer(file, delimiter=';')

                    writer.writerow([f'Missing Headers in {file1_path}'])
                    writer.writerows([[header] for header in missing_file1_headers])

                    writer.writerow([])

                    writer.writerow([f'Missing Data Values in {file1_path}'])
                    writer.writerows([[value] for value in missing_file1_data])

            else:
                messagebox.showwarning("Warning", f"The directory '{path}' does not contain two files.")
        else:
            messagebox.showwarning("Warning", f"The directory path '{path}' does not exist.")
    print("All completed")
    exit(0)


def write_results(output_file, missing_file1_headers, missing_file2_headers, missing_file1_data, missing_file2_data):
    # Prepare the data for writing
    rows = [
        ["Missing File 1 Headers:"] + missing_file1_headers,
        ["Missing File 2 Headers:"] + missing_file2_headers,
        ["Missing File 1 Data:"] + missing_file1_data,
        ["Missing File 2 Data:"] + missing_file2_data
    ]

    # Write the data to the output file
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        writer.writerows(rows)


# Create the main window
window = tk.Tk()
window.title("Bulk File Comparison")
window.geometry("300x200")

# Create the file selection label and entry
label_file1 = tk.Label(window, text="Select Excel File:")
label_file1.pack()
entry_file1 = tk.Entry(window, width=50)
entry_file1.pack()
button_file1 = tk.Button(window, text="Browse", command=lambda: browse_file(entry_file1))
button_file1.pack()

# Create the output file selection label and entry
label_output = tk.Label(window, text="Output File:")
label_output.pack()
entry_output = tk.Entry(window, width=50)
entry_output.pack()
button_output = tk.Button(window, text="Browse", command=lambda: browse_output_file(entry_output))
button_output.pack()

# Create the process button
button_process = tk.Button(window, text="Process Files", command=process_files)
button_process.pack()

# Run the main window loop
window.mainloop()
