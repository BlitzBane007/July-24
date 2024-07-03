import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import os
import pandas as pd
import csv
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import messagebox
from tqdm import tqdm


# Function to read the URL and download path from the Excel file
def read_data_from_excel(file_path):
    data = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    efs_url_column = headers.get("EFS Download URL ZIP")
    hfi_url_column = headers.get("hFI Download URL")
    path_column = headers.get("Path")
    if efs_url_column and path_column and hfi_url_column:
        for row in range(2, sheet.max_row + 1):
            efs_url_cell = sheet.cell(row=row, column=efs_url_column)
            hfi_url_cell = sheet.cell(row=row, column=hfi_url_column)
            path_cell = sheet.cell(row=row, column=path_column)
            data.append((efs_url_cell.value, hfi_url_cell.value, path_cell.value))

    return data


def has_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for item in os.listdir(directory_path):
        item_path = os.path.join(directory_path, item)
        if os.path.isdir(item_path):
            return True
    return False


def delete_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for subdirectory in os.listdir(directory_path):
        if os.path.isdir(subdirectory):
            os.system(f'rd /s /q "{subdirectory}"')


def create_subdir(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    path_col = "Path"

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    folder_name = headers.get("Folder Name")
    if folder_name:
        for row in range(2, sheet.max_row + 1):
            folder_cell = sheet.cell(row=row, column=folder_name)
            directory_name = folder_cell.value
            command_run = "mkdir "
            concat_str = " ".join([command_run, directory_name])
            subprocess.run(concat_str, shell=True)
    fetch_path = "dir /ad /b /s"
    fetch_output = subprocess.run(fetch_path, shell=True, capture_output=True)
    # Decode the output and split it into lines
    output_lines = fetch_output.stdout.decode().splitlines()

    # Strip any leading/trailing whitespace from each line
    output_lines = [line.strip() for line in output_lines]

    # Join the lines back together with line breaks
    column_index = None
    for column in sheet.iter_cols(min_row=1, max_row=1):
        if column[0].value == path_col:
            column_index = column[0].column_letter
            break
    # Write the output to the specific column
    if column_index is not None:
        for index, output_line in enumerate(output_lines):
            row_index = index + 2
            sheet[column_index + str(row_index)] = output_line
        wb.save(file_path)
    else:
        print(f"Column '{path_col}' not found.")


# Function to add parameter "&uncompressed=true" to the URL
def add_parameter_to_url(url):
    if url and "&uncompressed=true" not in url:
        if "?" in url:
            url += "&uncompressed=true"
        else:
            url += "?uncompressed=true"
    return url


# Function to download file using Google Chrome
def download_file(url, download_path, file_name):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        },
    )

    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(3)
    # Wait for the download to complete
    while True:
        download_dir = download_path
        files = os.listdir(download_dir)
        file_pathx = os.path.join(download_dir, files[0])
        if file_pathx.lower().endswith('.csv') or file_pathx.lower().endswith('.xlsx'):
            break
        time.sleep(1)  # Adjust the sleep time as needed
    driver.quit()
    # Rename the downloaded file
    files = os.listdir(download_path)
    file_path = os.path.join(download_path, files[0])
    new_file_path = os.path.join(download_path, file_name)
    os.rename(file_path, new_file_path)


# Main function to automate the task
def automate_task():
    # GUI: Interface to select the input Excel file
    print("Initialising STAGE 1")
    print("---------------------")
    file_path = entry_file1.get()

    if file_path:

        try:
            if has_subdirectories(file_path):
                delete_subdirectories(file_path)
                create_subdir(file_path)
            else:
                create_subdir(file_path)
        except OSError as e:
            print(f"Error executing .bat file: {e}")

        # Read the URL and download path from the Excel file
        data = read_data_from_excel(file_path)
        efs_file = "ZEFS.csv"
        hfi_file = "AhFI.csv"

        # Process each URL and download the file
        for efs_url, hfi_url, download_path in tqdm(data):
            modified_url = add_parameter_to_url(efs_url)
            download_file(modified_url, download_path, efs_file)
            download_file(hfi_url, download_path, hfi_file)
    else:
        print("No file selected.")


def read_csv_file(filename, sep):
    if not isinstance(sep, str):
        sep = ";"
    rows = []
    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file, delimiter=sep)
        for row in reader:
            rows.append(row)
    return rows


def read_xlsx_file(filename):
    rows = []
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows


def find_missing_headers(file1_headers, file2_headers):
    missing_file1 = set(file2_headers) - set(file1_headers)
    missing_file2 = set(file1_headers) - set(file2_headers)
    return missing_file1, missing_file2


def find_column_index(header_row, column_name):
    return header_row.index(column_name) if column_name in header_row else -1


def find_missing_data(file1_data, file2_data, column_index1, column_index2):
    file1_values = set(row[column_index1] for row in tqdm(file1_data[1:], desc='Processing File 1', unit='row',
                                                          ncols=80)) if column_index1 >= 0 else set()
    file2_values = set(row[column_index2] for row in tqdm(file2_data[1:], desc='Processing File 2', unit='row',
                                                          ncols=80)) if column_index2 >= 0 else set()
    missing_file1 = file2_values - file1_values
    missing_file2 = file1_values - file2_values
    return missing_file1, missing_file2


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def browse_output_file(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[('CSV Files', '*.csv')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def process_tree():
    # Run the automation task
    automate_task()
    process_files()


def process_files():
    # Get file path
    print("Initialising STAGE 2")
    print("---------------------")
    file_path = entry_file1.get()

    # Check if file is selected
    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return

    # Read Excel file
    try:
        df = pd.read_excel(file_path)
    except OSError:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return

    # Get the directory path values
    try:
        paths = df["Path"].tolist()
        delimiters = df["Delimiter"].tolist()
        # Check if any delimiter is blank and replace with ";"
        for i in range(len(delimiters)):
            if not delimiters[i]:
                delimiters[i] = ";"
    except KeyError:
        messagebox.showerror("Error", "The 'path' column does not exist in the Excel file.")
        return

    # Iterate over the paths
    for path, delimiter in zip(paths, delimiters):
        # Check if the path is valid
        if os.path.exists(path):
            # Get the list of files in the directory
            files = os.listdir(path)
            # Assuming the first two files in the directory are file 1 and file 2
            if len(files) >= 2:
                file_pathx = os.path.join(path, files[0])
                file_pathy = os.path.join(path, files[1])
                if file_pathx.lower().endswith('.csv') and file_pathy.lower().endswith('.xlsx'):
                    file1_path = os.path.join(path, files[0])
                    file1_data = read_csv_file(file1_path, delimiter)
                    file2_path = os.path.join(path, files[1])
                    file2_data = read_xlsx_file(file2_path)
                elif file_pathx.lower().endswith('.xlsx') and file_pathy.lower().endswith('.csv'):
                    file1_path = os.path.join(path, files[1])
                    file1_data = read_csv_file(file1_path, delimiter.value)
                    file2_path = os.path.join(path, files[0])
                    file2_data = read_xlsx_file(file2_path)
                elif file_pathx.lower().endswith('.csv') and file_pathy.lower().endswith('.csv'):
                    file1_path = os.path.join(path, files[1])
                    file1_data = read_csv_file(file1_path, delimiter)
                    file2_path = os.path.join(path, files[0])
                    file2_data = read_csv_file(file2_path, delimiter)
                else:
                    print(path)
                    continue

                # Check if headers are present
                if not file1_data or not file2_data:
                    messagebox.showwarning("Warning", "One or both files are empty.")
                    continue

                file1_headers = file1_data[0]
                file2_headers = file2_data[0]

                # Find missing headers
                missing_file1_headers, missing_file2_headers = find_missing_headers(file1_headers, file2_headers)

                # Find missing data
                column_index1 = find_column_index(file1_headers, "OFST020000")
                column_index2 = find_column_index(file2_headers, "OFST020000")
                missing_file1_data, missing_file2_data = find_missing_data(file1_data, file2_data, column_index1,
                                                                           column_index2)

                # Write the results to an output file

                output_path = entry_output.get()
                with open(output_path, 'a', encoding='utf-8', newline='') as file:
                    writer = csv.writer(file, delimiter=';')

                    writer.writerow([f'Missing Headers in {file1_path}'])
                    writer.writerows([[header] for header in missing_file1_headers])

                    writer.writerow([])

                    writer.writerow([f'Missing Data Values in {file1_path}'])
                    writer.writerows([[value] for value in missing_file1_data])

            else:
                messagebox.showwarning("Warning", f"The directory '{path}' does not contain two files.")
        else:
            messagebox.showwarning("Warning", f"The directory path '{path}' does not exist.")
    print("All completed")
    exit(0)


window = tk.Tk()
window.title("Bulk File Comparison")
window.geometry("300x200")

# Create the file selection label and entry
label_file1 = tk.Label(window, text="Select Excel File:")
label_file1.pack()
entry_file1 = tk.Entry(window, width=50)
entry_file1.pack()
button_file1 = tk.Button(window, text="Browse", command=lambda: browse_file(entry_file1))
button_file1.pack()

# Create the output file selection label and entry
label_output = tk.Label(window, text="Output File:")
label_output.pack()
entry_output = tk.Entry(window, width=50)
entry_output.pack()
button_output = tk.Button(window, text="Browse", command=lambda: browse_output_file(entry_output))
button_output.pack()

# Create the process button
button_process = tk.Button(window, text="Process Files", command=process_tree)
button_process.pack()

# Run the main window loop
window.mainloop()
