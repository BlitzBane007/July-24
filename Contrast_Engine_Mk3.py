import csv
import time
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from tqdm import tqdm
import threading

OFST020000= 'OFST020000'
def find_headers(file):
    headers_candidates = [next(file) for _ in range(2)]
    for i, candidate in enumerate(headers_candidates):
        if OFST020000 in candidate:  # Check if primary_k value is in this row
            return i, candidate
    raise ValueError("Headers not found in the first two lines of the file.")

def read_csv_file(filename, delimiter):
    rows = []
    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            rows.append(row)
    return rows

def read_xlsx_file(filename):
    rows = []
    wb = load_workbook(filename)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows

def select_efs_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("XLSX Files", "*.xlsx")])
    efs_file_label.config(text=file_path)

def select_hfi_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("XLSX Files", "*.xlsx")])
    hfi_file_label.config(text=file_path)

def compare_data():
    efs_file_path = efs_file_label.cget("text")
    hfi_file_path = hfi_file_label.cget("text")
    if not efs_file_path or not hfi_file_path:
        messagebox.showerror("Error", "Please select both EFS and HFI files.")
        return

    output_file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
    if not output_file_path:
        return

    try:
        delimiter = delimiter_entry.get()

        if efs_file_path.endswith('.csv'):
            efs_data = read_csv_file(efs_file_path, delimiter)
        else:
            efs_data = read_xlsx_file(efs_file_path)

        if hfi_file_path.endswith('.csv'):
            hfi_data = read_csv_file(hfi_file_path, delimiter)
        else:
            hfi_data = read_xlsx_file(hfi_file_path)

        efs_line, efs_headers = find_headers(iter(efs_data))
        hfi_line, hfi_headers = find_headers(iter(hfi_data))

        # if efs_line == 1:  # If headers are in the second row, remove the first line from efs_data
        #     efs_data = efs_data[1:]
        # if hfi_line == 1:  # If headers are in the second row, remove the first line from hfi_data
        #     hfi_data = hfi_data[1:]

        output_data = [["isin", "header", "efs", "hfi"]]

        total_headers = len(efs_headers) - 1  # Exclude the "primary_k" header

        # # Create the progress bar
        # progress_bar = tk.Label(window, text="Processing Headers...")
        # progress_bar.pack(pady=10)

        for header_index, efs_header in tqdm(enumerate(efs_headers), desc="Checking headers", unit="Headers"):
            if efs_header != OFST020000:
                if efs_header in hfi_headers:
                    for efs_row in tqdm(efs_data, desc="Checking rows", unit="Row"):
                        efs_isin = efs_row[efs_headers.index(OFST020000)]
                        for hfi_row in hfi_data:
                            hfi_isin = hfi_row[hfi_headers.index(OFST020000)]
                            if efs_isin == hfi_isin:
                                efs_value = efs_row[efs_headers.index(efs_header)]
                                hfi_value = hfi_row[hfi_headers.index(efs_header)]
                                if efs_value != hfi_value:
                                    # print('not match', efs_value, hfi_value)
                                    output_data.append([efs_isin, efs_header, efs_value, hfi_value])
                                    # time.sleep(2)


            # progress_bar.config(text=f"Processing Header {header_index + 1}/{total_headers}")
            # window.update()

        # progress_bar.destroy()

        with open(output_file_path, "w", newline="", encoding="utf-8") as output_file:
            writer = csv.writer(output_file)
            writer.writerows(output_data)

        messagebox.showinfo("Done", "Compare completed!")
    except FileNotFoundError:
        messagebox.showerror("Error", "File not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


# Create the GUI window
window = tk.Tk()
window.title("CSV Comparison Tool")

# Create the "Select EFS File" button
efs_button = tk.Button(window, text="Select EFS File", command=select_efs_file)
efs_button.pack(pady=10)

# Create a label to display the selected EFS file path
efs_file_label = tk.Label(window, text="")
efs_file_label.pack()

# Create the "Select HFI File" button
hfi_button = tk.Button(window, text="Select HFI File", command=select_hfi_file)
hfi_button.pack(pady=10)

# Create a label to display the selected HFI file path
hfi_file_label = tk.Label(window, text="")
hfi_file_label.pack()

# Create a label and entry for delimiter
delimiter_label = tk.Label(window, text="Delimiter:")
delimiter_label.pack()
delimiter_entry = tk.Entry(window)
delimiter_entry.insert(tk.END, ";")
delimiter_entry.pack()

# Create the "Compare Data" button
compare_button = tk.Button(window, text="Compare Data", command=compare_data)
compare_button.pack(pady=10)

# Start the GUI event loop
window.mainloop()