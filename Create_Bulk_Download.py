import time
from tkinter import Tk
from tkinter import filedialog
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from tqdm import tqdm
import subprocess


# Function to read the URL and download path from the Excel file
def read_data_from_excel(file_path):
    data = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    efs_url_column = headers.get("EFS Download URL ZIP")
    hfi_url_column = headers.get("hFI Download URL")
    path_column = headers.get("Path")
    if efs_url_column and path_column and hfi_url_column:
        for row in range(2, sheet.max_row + 1):
            efs_url_cell = sheet.cell(row=row, column=efs_url_column)
            hfi_url_cell = sheet.cell(row=row, column=hfi_url_column)
            path_cell = sheet.cell(row=row, column=path_column)
            data.append((efs_url_cell.value, hfi_url_cell.value, path_cell.value))

    return data


# Function to add parameter "&uncompressed=true" to the URL
def add_parameter_to_url(url):
    if url and "&uncompressed=true" not in url:
        if "?" in url:
            url += "&uncompressed=true"
        else:
            url += "?uncompressed=true"
    return url


# Function to download file using Google Chrome
def download_file(url, download_path):
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        },
    )

    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(3)

    # Get the initial file count in the download directory

    # Wait for the download to complete
    while True:
        download_dir = download_path
        files = os.listdir(download_dir)
        file_pathx = os.path.join(download_dir, files[0])
        if file_pathx.lower().endswith('.csv') or file_pathx.lower().endswith('.xlsx'):
            break
        time.sleep(1)  # Adjust the sleep time as needed

    # Add a delay before closing the Chrome window
    time.sleep(1)

    driver.quit()


# Main function to automate the task
def automate_task():
    # GUI: Interface to select the input Excel file
    root = Tk()
    root.withdraw()  # Hide the Tkinter window
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        try:
            # Change the working directory to the specified folder location
            directory_path = os.path.dirname(file_path)
            os.chdir(directory_path)

            # Execute the .bat file using the subprocess.call() function
            subprocess.call("create.bat", shell=True)
        except OSError as e:
            print(f"Error executing .bat file: {e}")
        # Read the URL and download path from the Excel file
        data = read_data_from_excel(file_path)

        # Process each URL and download the file
        for efs_url, hfi_url, download_path in tqdm(data):
            modified_url = add_parameter_to_url(efs_url)
            download_file(modified_url, download_path)
            download_file(hfi_url, download_path)
    else:
        print("No file selected.")


# Run the automation task
automate_task()
