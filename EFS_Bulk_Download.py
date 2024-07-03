import time
from tkinter import Tk
from tkinter import filedialog

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


# Function to read the URL and download path from the Excel file
def read_data_from_excel(file_path):
    data = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    url_column = headers.get("Download URL ZIP")
    path_column = headers.get("Path")
    if url_column and path_column:
        for row in range(2, sheet.max_row + 1):
            url_cell = sheet.cell(row=row, column=url_column)
            path_cell = sheet.cell(row=row, column=path_column)
            data.append((url_cell.value, path_cell.value))

    return data


# Function to add parameter "&uncompressed=true" to the URL
def add_parameter_to_url(url):
    if url and "&uncompressed=true" not in url:
        if "?" in url:
            url += "&uncompressed=true"
        else:
            url += "?uncompressed=true"
    return url


# Function to download file using Google Chrome
def download_file(url, download_path):
    options = Options()
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": download_path,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        },
    )

    driver = webdriver.Chrome(options=options)
    driver.get(url)

    # Wait for the download to complete
    WebDriverWait(driver, 60).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='downloadSpinner']")))

    # Add a delay before closing the Chrome window
    time.sleep(5)

    driver.quit()


# Main function to automate the task
def automate_task():
    # GUI: Interface to select the input Excel file
    root = Tk()
    root.withdraw()  # Hide the Tkinter window
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        # Read the URL and download path from the Excel file
        data = read_data_from_excel(file_path)

        # Process each URL and download the file
        for url, download_path in data:
            modified_url = add_parameter_to_url(url)
            download_file(modified_url, download_path)
    else:
        print("No file selected.")


# Run the automation task
automate_task()
