"""
This program helps with picking the download URL and folder for a specific feed and then downloading the latest file.
The files are then compared and the missing ISIN and headers in the files are listed.
NOTE: The primary Excel file in correct format is required.
Added functions:
Ability to map the missing ISIN to their citicodes using Enigma Feed.
The citicodes are then used to show the blocking filters for the files.
Refined the logic to check the latest files and also enigma.
Added ability to download the ZIP files instead of the uncompressed versions.
NOTE: The hFI download link should point to the zip file.
"""

import datetime
import os
import subprocess
import time
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import dask.dataframe as dd
import openpyxl
import pandas as pd
from dask.diagnostics import ProgressBar
from urllib.parse import urlparse
import json
import requests
from tqdm import tqdm
import zipfile
import webbrowser
from dask import delayed
import re


# Function to read the URL and download path from the Excel file
def read_data_from_excel(file_path):
    data = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    efs_url_column = headers.get("EFS")
    hfi_url_column = headers.get("HFI")
    path_column = headers.get("Path")
    if efs_url_column and path_column and hfi_url_column:
        for row in range(2, sheet.max_row + 1):
            efs_url_cell = sheet.cell(row=row, column=efs_url_column)
            hfi_url_cell = sheet.cell(row=row, column=hfi_url_column)
            path_cell = sheet.cell(row=row, column=path_column)
            data.append((efs_url_cell.value, hfi_url_cell.value, path_cell.value))

    return data


def has_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for item in os.listdir(directory_path):
        item_path = os.path.join(directory_path, item)
        if os.path.isdir(item_path):
            return True
    return False


def download_enigma(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    enig_file = "Enigma.csv"
    url = "https://datafeeds.fefundinfo.com/api/v1/Feeds/22a151f4-7937-4b3f-b3bd-1f440d14e62e/download?token=f26136d9" \
          "-21a8-4d62-9694-e11c1a14a40b"
    download_and_extract_zip(url, directory_path, enig_file)


def delete_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for subdirectory in os.listdir(directory_path):
        if os.path.isdir(subdirectory):
            os.system(f'rd /s /q "{subdirectory}"')


def create_subdir(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    path_col = "Path"

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    folder_name = headers.get("FolderName")
    if folder_name:
        for row in range(2, sheet.max_row + 1):
            folder_cell = sheet.cell(row=row, column=folder_name)
            directory_name = folder_cell.value
            command_run = "mkdir "
            concat_str = " ".join([command_run, directory_name])
            subprocess.run(concat_str, shell=True)
    fetch_path = "dir /ad /b /s"
    fetch_output = subprocess.run(fetch_path, shell=True, capture_output=True)
    # Decode the output and split it into lines
    output_lines = fetch_output.stdout.decode().splitlines()

    # Strip any leading/trailing whitespace from each line
    output_lines = [line.strip() for line in output_lines]

    # Join the lines back together with line breaks
    column_index = None
    for column in sheet.iter_cols(min_row=1, max_row=1):
        if column[0].value == path_col:
            column_index = column[0].column_letter
            break
    # Write the output to the specific column
    if column_index is not None:
        for index, output_line in enumerate(output_lines):
            row_index = index + 2
            sheet[column_index + str(row_index)] = output_line
        wb.save(file_path)
    else:
        print(f"Column '{path_col}' not found.")


def downloadlatest_loop(file_path):
    data = read_data_from_excel(file_path)
    len(data)
    count = len(data)
    print(f"Downloading ", count, "files!")
    efs_file = "ZEFS.csv"
    for efs_url, hfi_url, download_path in data:
        print(f"Downloading EFS file to", download_path)
        download_and_extract_zip(efs_url, download_path, efs_file)


def download_and_extract_zip(url, path, new_filename):
    # Extract the filename from the URL
    parsed_url = urlparse(url)
    filename = os.path.basename(parsed_url.path)

    # Download the zip file
    response = requests.get(url, stream=True)

    # Get the total file size in bytes
    total_size = int(response.headers.get('content-length', 0))

    # Save the downloaded zip file
    zip_filename = os.path.join(path, filename)
    with open(zip_filename, 'wb') as file:
        # Create a progress bar using tqdm
        progress_bar = tqdm(total=total_size, unit='B', unit_scale=True)

        # Iterate over the response content in chunks
        for data in response.iter_content(chunk_size=4096):
            # Write the chunk to the file
            file.write(data)

            # Update the progress bar
            progress_bar.update(len(data))

        # Close the progress bar
        progress_bar.close()

    # Extract the contents of the zip file to the specified path
    with zipfile.ZipFile(zip_filename, 'r') as zip_ref:
        zip_ref.extractall(path)

    # Rename the extracted file
    extracted_files = zip_ref.namelist()
    for extracted_file in extracted_files:
        extracted_file_path = os.path.join(path, extracted_file)
        new_file_path = os.path.join(path, new_filename)
        os.rename(extracted_file_path, new_file_path)
        break  # Only rename the first file

    # Delete the original zip file
    os.remove(zip_filename)


def download_loop(file_path):
    data = read_data_from_excel(file_path)
    count = len(data) * 2
    print(f"Downloading ", count, "files!")
    efs_file = "ZEFS.csv"
    hfi_file = "AhFI.csv"
    for efs_url, hfi_url, download_path in data:
        print(f"Downloading EFS file to", download_path)
        download_and_extract_zip(efs_url, download_path, efs_file)
        print(f"Downloading hFI file to", download_path)
        download_and_extract_zip(hfi_url, download_path, hfi_file)


# Main function to automate the task
def automate_task():
    # GUI: Interface to select the input Excel file
    print("Initialising STAGE 1")
    print("---------------------")
    start = time.time()

    file_path = entry_file1.get()

    if file_path:
        try:
            print("CHECKING FOLDER INTEGRITY")
            # Read the URL and download path from the Excel file

            if has_subdirectories(file_path):
                print("FOLDER INTEGRITY - PASS")
                print('--------------------')
                print("CHECKING ALL FILE INTEGRITY")
                flag = check_latest(file_path)
                if flag == 1:
                    print("ENIGMA INTEGRITY - LATEST")
                    print("hFI FILE INTEGRITY - LATEST")
                    print("STARTING EFS DOWNLOAD")
                    # Process each EFS URL and download the LATEST file
                    downloadlatest_loop(file_path)
                elif flag == 2:
                    print("ENIGMA - OUT OF DATE")
                    print("UPDATING ENIGMA")
                    download_enigma(file_path)
                    print("ENIGMA UPDATED")
                    print('--------------------')
                    print("STARTING EFS DOWNLOAD")
                    downloadlatest_loop(file_path)
                else:
                    print('FILE INTEGRITY : OBSOLETE')
                    delete_subdirectories(file_path)
                    create_subdir(file_path)
                    print("UPDATING ENIGMA")
                    # download_enigma(file_path)
                    print("UPDATING DATA FILES")
                    download_loop(file_path)
            else:
                print('FOLDER INTEGRITY : NON-EXISTENT')
                print('STARTING FRESH DOWNLOAD')
                print('--------------------')
                create_subdir(file_path)
                print("UPDATING ENIGMA")
                download_enigma(file_path)
                print("UPDATING DATA FILES")
                download_loop(file_path)
        except OSError as e:
            print(f"Error executing .bat file: {e}")
    else:
        print('file path issue')
    if not Flag:
        print(f"Download completed in", time.time() - start)


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def browse_output_file(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[('CSV Files', '*.csv')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def process_tree():
    global Flag
    # Check the state of the checkboxes
    run_automate = var_automate.get()
    run_process = var_process.get()
    Flag = run_process

    # Create and start a new thread for running the UI
    if run_automate:
        automate_task()

    # Run the process_files function in a separate thread
    if run_process:
        process_files()


def check_latest(file_path):
    data_quality = []
    file_name = 'AhFI.csv'
    file_enig = 'Enigma.csv'
    data2 = 1
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    dir_files = os.listdir(directory_path)
    for dir_file in dir_files:
        if dir_file == file_enig:
            print("FOUND ENIGMA - CHECKING INTEGRITY")
            cur_path = os.path.join(directory_path, dir_file)
            modified_timestamp = os.path.getmtime(cur_path)
            modified_datetime = datetime.datetime.fromtimestamp(modified_timestamp)
            current_datetime = datetime.datetime.now()
            time_difference = current_datetime - modified_datetime
            if time_difference.days > 7:
                print("ENIGMA IS OUT OF DATE")
                data2 = 1
                enig_path = os.path.join(directory_path, file_enig)
                if os.path.exists(enig_path):
                    os.remove(enig_path)
            else:
                data2 = 0
    try:
        df = pd.read_excel(file_path)
    except OSError:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return
    paths = df["Path"].tolist()
    for path in paths:
        if str(path) != "nan":
            if os.path.exists(path):
                # Get the list of files in the directory
                files = os.listdir(path)
                for file in files:
                    if file == file_name:
                        cur_path = os.path.join(path, file_name)
                        modified_timestamp = os.path.getmtime(cur_path)
                        modified_datetime = datetime.datetime.fromtimestamp(modified_timestamp)
                        current_datetime = datetime.datetime.now()
                        time_difference = current_datetime - modified_datetime
                        if time_difference.days < 2:
                            data_quality.append(0)
                            efs_path = os.path.join(path, "ZEFS.csv")
                            if os.path.exists(efs_path):
                                os.remove(efs_path)
                            else:
                                continue
                        else:
                            return False
    for data1 in data_quality:
        if data1 == 0 and data2 == 0:
            return True
        elif data1 == 0 and data2 == 1:
            return 2
        else:
            return False


def extract_api_from_url(url):
    parsed_url = urlparse(url)
    path_segments = parsed_url.path.split('/')
    api = path_segments[-2]  # Extract the second-to-last segment

    return api


# Filter the DataFrame based on the value of OFST020000 and retrieve the corresponding value of OFST900174
def find_associated_value(df, value_ofst020000):
    filtered_df = df[df['OFST020000'] == value_ofst020000]
    associated_value = None
    if not filtered_df.empty:
        associated_value = filtered_df['OFST900174'].iloc[0]  # Assuming there's only one associated value
    return associated_value


def call_api(api_key, citicode):
    url = f'https://datafeeds.fefundinfo.com/api/data/filtercheck/{api_key}?citiCode={citicode}'
    response = requests.get(url)
    data = json.loads(response.text)
    blocking_filters_tag = data['blockingFilters']
    return str(blocking_filters_tag)


def open_efs_feed():
    print("Opening EFS feed page")
    file_path = entry_file1.get()
    data = read_data_from_excel(file_path)
    for efs_url, hfi_url, download_path in data:
        parsed_url = urlparse(efs_url)
        path_segments = parsed_url.path.split('/')
        api = path_segments[-2]  # Extract the second-to-last segment
        feed_url = f"https://datafeeds.fefundinfo.com/feeds/feeds?id={api}"
        webbrowser.open_new_tab(feed_url)


def open_hfi_feed():
    print("Opening hFI feed page")
    file_path = entry_file1.get()
    data = read_data_from_excel(file_path)
    for efs_url, hfi_url, download_path in data:
        parsed_url = urlparse(hfi_url)
        path_segments = parsed_url.path.split('/')
        api = path_segments[-2]  # Extract the second-to-last segment
        feed_url = f"https://datafeed.fundinfo.com/Feeds/DataFeed/Edit/{api}"
        webbrowser.open_new_tab(feed_url)


@delayed
def read_excel_with_pandas(file_path):
    return pd.read_excel(file_path, engine='openpyxl')


def process_files():
    # Get file path
    global primary_k
    output_data = []
    TCOUNT = 0
    head3 = "Blocking Filters"
    head4 = "Citicode"
    print("Initialising STAGE 2")
    print("---------------------")
    start_time = time.time()
    file_path = entry_file1.get()
    fc_count = entry_filter.get()

    # Check if file is selected
    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return

    # Read Excel file
    try:
        df = pd.read_excel(file_path)
    except OSError:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return

    # Get the directory path values

    paths = df["Path"].tolist()
    delimiters = df["Delimiter"].tolist()
    urls = df["EFS"].tolist()

    # Check if any delimiter is blank and replace with ";"
    for i in range(len(delimiters)):
        if str(delimiters[i]) == "nan":
            delimiters[i] = ";"

    file_enig = 'Enigma.csv'
    edir_path = os.path.dirname(file_path)
    enigma_path = os.path.join(edir_path, file_enig)
    if os.path.exists(enigma_path):
        df_enig = pd.read_csv(enigma_path, delimiter=';')
    else:
        messagebox.showerror("Error", "Enigma not found - Downloading Enigma!")
        download_enigma(file_path)
        print('Enigma Integrity - PASS')
        df_enig = pd.read_csv(enigma_path, delimiter=';')

    # Iterate over the paths
    for path, delimiter, url in zip(paths, delimiters, urls):
        api = extract_api_from_url(url)
        # Check if the path is valid
        if os.path.exists(path):
            # Get the list of files in the directory
            files = os.listdir(path)
            # Assuming the first two files in the directory are file 1 and file 2
            if len(files) >= 2:
                file1_path = os.path.join(path, files[1])
                file2_path = os.path.join(path, files[0])
                blocks = []
                citi = []

                print("Initialising Data Loading...")
                print(f'File :', file1_path)
                if file1_path.lower().endswith('.csv'):
                    df_efs = dd.read_csv(file1_path, header=0, sep=delimiter, low_memory=False, assume_missing=True,
                                         blocksize=1e9,
                                         skiprows=[1], dtype=str)

                if file2_path.lower().endswith('.csv'):
                    df_hfi = dd.read_csv(file2_path, header=0, sep=delimiter, low_memory=False, assume_missing=True,
                                         blocksize=1e9,
                                         skiprows=[1], dtype=str)

                if file1_path.lower().endswith('.xlsx'):
                    df_efs = dd.from_delayed(read_excel_with_pandas(file1_path))

                if file2_path.lower().endswith('.xlsx'):
                    df_hfi = dd.from_delayed(read_excel_with_pandas(file2_path))

                df_efs = df_efs.compute()

                df_hfi = df_hfi.compute()

                # Get the union of unique headers
                union_headers = pd.unique(df_efs.columns.union(df_hfi.columns))
                # Convert the union_headers to a list if needed
                headers = [re.sub(r'[()-]', '', header) for header in union_headers]

                print('--------------')

                # Get headers missing in df_efs
                missing_headers = pd.Series(list(set(headers) - set(df_efs.columns)))
                print('Headers missing in efs:', len(missing_headers))
                head1 = f'Headers missing in {api}:'
                missing_headers_df = pd.DataFrame({head1: missing_headers})
                output_data.append(missing_headers_df)

                # Get ISIN missing in df_efs
                missing_isins = pd.Series(list(set(df_hfi[primary_k].unique()) - set(df_efs[primary_k].unique())))
                print('ISIN missing in efs:', len(missing_isins))
                TCOUNT += len(missing_isins)
                head2 = f'ISIN missing in {api}'
                missing_isin_df = pd.DataFrame({head2: missing_isins})
                output_data.append(missing_isin_df)
                if var_filter.get():
                    print("FILTER CHECK INITIATED")
                    counter = 0
                    for missing_isin in tqdm(missing_isins):
                        counter += 1
                        if counter > int(fc_count):
                            break
                        citicode = find_associated_value(df_enig, missing_isin)
                        if str(citicode) == "None":
                            citi.append("No citicode on DB")
                            blocks.append("NA")
                        else:
                            citi.append(str(citicode))
                            blocker = call_api(api, str(citicode))
                            blocks.append(blocker)

                    citi_df = pd.DataFrame({head4: citi})
                    blocker_df = pd.DataFrame({head3: blocks})
                    output_data.append(citi_df)
                    output_data.append(blocker_df)
                    print(citi)
                    print(blocks)

                print("-----------------")

            else:
                messagebox.showwarning("Warning", f"The directory '{path}' does not contain two files.")
        else:
            messagebox.showwarning("Warning", f"The directory path '{path}' does not exist.")

    print("Initialising STAGE 4 - Write data to CSV file")
    output_path = entry_output.get()
    output_data = pd.concat(output_data, axis=1)
    output_data.to_csv(output_path, index=False, mode='w')
    print("--------------")
    print('Time taken:', time.time() - start_time)
    print("All completed")
    print(TCOUNT)


window = tk.Tk()
window.title("Bulk File Comparison")
window.geometry("350x350")
ProgressBar().register()
primary_k = 'OFST020000'
Flag = 0

# Create the file selection label and entry
label_file1 = tk.Label(window, text="Select Excel File:")
label_file1.pack()
entry_file1 = tk.Entry(window, width=50)
entry_file1.pack()
button_file1 = tk.Button(window, text="Browse", command=lambda: browse_file(entry_file1))
button_file1.pack()

# Create the output file selection label and entry
label_output = tk.Label(window, text="Output File:")
label_output.pack()
entry_output = tk.Entry(window, width=50)
entry_output.pack()
button_output = tk.Button(window, text="Browse", command=lambda: browse_output_file(entry_output))
button_output.pack()

# Create the checkboxes

var_automate = tk.IntVar()
checkbox_automate = tk.Checkbutton(window, text="DOWNLOAD FILES", variable=var_automate)
checkbox_automate.pack()

var_process = tk.IntVar()
checkbox_process = tk.Checkbutton(window, text="COMPARE DATA", variable=var_process)
checkbox_process.pack()

var_filter = tk.IntVar()
checkbox_filter = tk.Checkbutton(window, text="FILTER CHECK", variable=var_filter)
checkbox_filter.pack()

# Create the process button
button_process = tk.Button(window, text="Process Files", command=process_tree)
button_process.pack()

# Create the efs button
button_EFS = tk.Button(window, text="Open EFS feed page", command=open_efs_feed)
button_EFS.pack()

# Create the hfi button
button_HFI = tk.Button(window, text="Open hFI feed page", command=open_hfi_feed)
button_HFI.pack()
# Create the output file selection label and entry
label_filter = tk.Label(window, text="Filtercheck count")
label_filter.pack()
entry_filter = tk.Entry(window, width=50)
entry_filter.pack()

# Run the main window loop
window.mainloop()
