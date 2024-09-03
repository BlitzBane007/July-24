import os
import requests
import json
import logging
import datetime
import sys
import openpyxl
import xml.etree.ElementTree as ET
import ast
import time
import uuid
import re
import urllib.parse
import random

TIMEOUT_DURATION = 30

logging.basicConfig(filename='recpterror.log', level=logging.ERROR, format='%(asctime)s:%(levelname)s:%(message)s')

SubId = ''
Perm_Cont = ''
DUMP = ''
TRACKER = ''
user_path = ''
proj_guid = ''
guid_list = []
recpt_token = ''
diss_token = ''
recpt_read_token = ''
diss_read_token = ''
flag = 0
linked_recpts = set()



def log_message(message):
    global user_path
    # Step 2: Write the message to the standard output
    print(str(message))
    # Step 3: Append the message to the log file
    # Step 2: Get the current date and time
    current_datetime = datetime.datetime.now()
    # Step 3: Format the date and time (optional)
    formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
    log_file_path = os.path.join(user_path, 'Output.log')
    with open(log_file_path, 'a') as log_file:
        log_file.write(formatted_datetime + ' - ' + str(message) + '\n')


def get_user():
    global user_path, DUMP, TRACKER
    user_path = input("Please enter the OneDrive Path to 'FSQtoEFS'\nUser Path:")
    log_message("Thank you!")
    DUMP = os.path.join(user_path, 'RecipientDump.xlsx')
    TRACKER = os.path.join(user_path, 'TRACKER.xlsx')
    log_message(f'Dump path built: {DUMP}')
    log_message(f'Tracker path built: {TRACKER}')


def Script_Run():
    global DUMP, TRACKER, guid_list, recpt_token, diss_token, proj_guid, recpt_read_token, diss_read_token, flag, linked_recpts

    run_count = 0
    counter = int(input("Please enter the number of iterations:"))
    if not 1 <= counter <= 1000:
        log_message("Number must be between 1 and 1000.")
        input("Hit ENTER to Exit!")
        sys.exit(1)

    wbt = openpyxl.load_workbook(filename=TRACKER, data_only=True)
    wst = wbt["Tracker"]

    wbd = openpyxl.load_workbook(filename=DUMP, data_only=True)
    wsd = wbd["DUMP"]

    def get_Diss_bearer_token():
        log_message("Getting Diss Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'fefundinfo-esf-dissemination-api-dissemination-mutate'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def get_Recpt_bearer_token():
        log_message("Getting Recpt Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'fefundinfo-esf-dissemination-api-recipient-mutate'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def get_Recpt_read_token():
        log_message("Getting Recpt Read Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'fefundinfo-esf-dissemination-api-recipient-read'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def get_Diss_read_token():
        log_message("Getting Diss Read Token!")
        url = 'https://auth.fefundinfo.com/connect/token'
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        data = {
            'client_id': 'EFS-migration-for-support',
            'client_secret': 'iz63fbucsQ9IEQKIC5eveeGpNlK8MfV',
            'grant_type': 'client_credentials',
            'scope': 'fefundinfo-esf-dissemination-api-dissemination-read'
        }
        try:
            response = requests.post(url, headers=headers, data=data, timeout=TIMEOUT_DURATION)
            response.raise_for_status()
            return response.json()['access_token']
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            input("Hit Enter to Exit")
            exit(1)

    def get_tracker_index_by_header(header):
        for col in wst.iter_cols(1, wst.max_column):
            if col[0].value == header:
                return col[0].column
        return None

    def get_dump_index_by_header(header):
        for col in wsd.iter_cols(1, wsd.max_column):
            if col[0].value == header:
                return col[0].column
        return None

    def sanitize_email(email):
        local_part = email.split('@')[0]
        sanitized_local_part = re.sub(r'\W+', '', local_part)
        return sanitized_local_part

    def Create_Recpt(payload):
        url = f"https://efs-dissemination.fefundinfo.com/recipient"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {recpt_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }
        data = payload

        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 201:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            create_status = data.get('payload', {}).get('message', '')
            guid = data["payload"].get('guid', '')
            print(f'{create_status} : {guid}')
            guid_list.append(guid)

        else:
            log_message(f"Create Recpt Request failed with status code {response.status_code}: {response.text}")

    def Create_Diss(payload):
        global proj_guid
        url = f"https://efs-dissemination.fefundinfo.com/feed"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }
        data = payload

        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 201:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            print(response_text)
            data = json.loads(response_text)
            create_status = data.get('payload', {}).get('message', '')
            proj_guid = data["payload"].get('guid', '')
            print(f'{create_status} : {proj_guid}')
        else:
            log_message(f"Create Diss Request failed with status code {response.status_code}: {response.text}")

    def Linking(payload, pro_guid, cont):
        url = f"https://efs-dissemination.fefundinfo.com/container/{cont}/project/{pro_guid}/subscribers"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }
        data = payload


        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 201:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            create_status = data.get('payload', {}).get('message', '')
            print(f'{create_status}')
        else:
            log_message(f"Linking Request failed with status code {response.status_code}: {response.text}")

    def Search_Recpt(email, perm_cont):

        def convert_email(email1):
            return urllib.parse.quote(email1)
        converted_email = convert_email(email)
        url = f"https://efs-dissemination.fefundinfo.com/recipients?Top=10&Skip=0&OrderBy=recipientName&Descending=false&Name={converted_email}&containerIds={perm_cont}"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {recpt_read_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            for recipient in data["payload"]["recipients"]:
                res_email = recipient['emailAddress']
                res_cont = recipient['containerId']
                guid = recipient['guid']
                if res_email == email and res_cont == perm_cont:
                    guid_list.append(guid)
                    print(f'{guid}')
                    return True
                else:
                    return False
        else:
            log_message(f"Linking Request failed with status code {response.status_code}: {response.text}")

    def Search_Diss(name, perm_cont):
        global proj_guid
        url = f"https://efs-dissemination.fefundinfo.com/dissemination?Top=9&Skip=0&OrderBy=projectName&Descending=false&Name={name}"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_read_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_text = json.dumps(response.json(), indent=4)  # Format JSON output
            data = json.loads(response_text)
            proj_guid = data["payload"]["projects"][0].get('guid', '')
            proj_cont = data["payload"]["projects"][0].get('containerId', '')
            if proj_cont == perm_cont:
                return True
            else:
                return False
        else:
            log_message(f"Linking Request failed with status code {response.status_code}: {response.text}")

    def Del_recpts(proj, perm_cont,list):
        url = f"https://efs-dissemination.fefundinfo.com/container/{perm_cont}/project/{proj}/subscriber"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        response = requests.delete(url, headers=headers, json=list)
        print(response.status_code)
        if response.status_code == 204:
            print("Recpts deleted")
            return True
        else:
            log_message(f"Deletion Request failed with status code {response.status_code}: {response.text}")

    def Get_recpt_count(proj_guid, perm_cont_id, route):
        global guid_list, linked_recpts
        log_message("Getting Recpt count Details")
        url = f"https://efs-dissemination.fefundinfo.com/container/{perm_cont_id}/project/{proj_guid}/subscriber?Top=1000"
        headers = {
            'accept': 'application/json',
            'Authorization': f'Bearer {diss_token}',  # Replace YOUR_TOKEN with your actual token
            'Content-Type': 'application/json',
        }

        try:
            response = requests.get(url, headers=headers)

            if response.status_code == 200:
                response_text = json.dumps(response.json(), indent=4)  # Format JSON output
                data = json.loads(response_text)
                recipient_guids = set(project['recipientGuid'] for project in data['payload']['projects'])
                if route == 0:
                    linked_recpts = recipient_guids
                if route == 1:
                    guid_list = [guid for guid in guid_list if guid not in recipient_guids]
            else:
                log_message(f"Request failed with status code {response.status_code}: {response.text}")
        except Exception as e:
            logging.exception("An error occurred in Bearer Token: %s", e)
            log_message(e)
            Swag_recpt_count = 0

    tracker_sub_id = "Sub_ID"
    tracker_perm_cont_id = "Perm_Cont_ID"
    tracker_feed_id = "Feed_ID"
    tracker_status = "Status"

    tracker_idx_sub_id = get_tracker_index_by_header(tracker_sub_id)
    tracker_idx_perm_cont_id = get_tracker_index_by_header(tracker_perm_cont_id)
    tracker_idx_feed_id = get_tracker_index_by_header(tracker_feed_id)
    tracker_idx_status = get_tracker_index_by_header(tracker_status)

    dump_sub_id = "SubscriptionID"
    dump_recpt_type = "RecipientType"
    dump_email = "Email_ID"

    dump_sub_id_idx = get_dump_index_by_header(dump_sub_id)
    dump_recpt_type_idx = get_dump_index_by_header(dump_recpt_type)
    dump_email_idx = get_dump_index_by_header(dump_email)

    recpt_token = get_Recpt_bearer_token()
    diss_token = get_Diss_bearer_token()
    recpt_read_token = get_Recpt_read_token()
    diss_read_token = get_Diss_read_token()

    for row in range(2, wst.max_row + 1):
        if run_count == counter:
            break
        track_status = wst.cell(row=row, column=tracker_idx_status)
        if track_status.value == 'Error':
            run_count += 1
            print("===============================")
            print("===============================")
            print(f'Run Count = {run_count}')
            Recipient_list = []
            guid_list = []
            recpt_count = 0
            Tracker_sub_id_value = wst.cell(row=row, column=tracker_idx_sub_id).value
            Tracker_perm_cont_id_value = wst.cell(row=row, column=tracker_idx_perm_cont_id).value
            Tracker_feed_id_value = wst.cell(row=row, column=tracker_idx_feed_id).value
            print(f'Processing Feed ID: {Tracker_feed_id_value}')
            for d_row in range(2, wsd.max_row + 1):
                dump_sub_id_value = wsd.cell(row=d_row, column=dump_sub_id_idx).value
                if dump_sub_id_value == Tracker_sub_id_value:
                    recpt_count += 1
                    dump_recpt_type_value = wsd.cell(row=d_row, column=dump_recpt_type_idx).value
                    dump_email_value = wsd.cell(row=d_row, column=dump_email_idx).value
                    username = sanitize_email(dump_email_value)
                    Recipient_list.append((username, dump_recpt_type_value, dump_email_value))
            print(f'Recipient Count = {recpt_count}')
            print("Creating Recipients")
            for user, type, email in Recipient_list:
                random_3_digit = random.randint(100, 999)
                user = f'MIG-FSQ{random_3_digit}-{user}'
                unzip = True
                if type == "MZIP":
                    unzip = False
                else:
                    unzip = True
                if type == "MAIL" or type == "MZIP":
                    recpt_payload = {
                        "recipientName": user,
                        "deliveryMethod": 3,
                        "emailAddress": email,
                        "shouldUnzipFile": unzip,
                        "embargoConsumerId": None,
                        "language": "",
                        "emailNotification": False,
                        "attachFile": True,
                        "containerId": Tracker_perm_cont_id_value
                    }
                    recpt_found = Search_Recpt(email, Tracker_perm_cont_id_value)
                    if not recpt_found:
                        print(f'Recipient : {email} not found in {Tracker_perm_cont_id_value}')
                        stop_flag = "y"
                        # stop_flag = input("Type x to exit")
                        if stop_flag == "x":
                            exit(0)
                        elif stop_flag == "y":
                            Create_Recpt(recpt_payload)
                else:
                    print(f'Sub ID: {Tracker_sub_id_value} has other type of diss recpt')
                    flag = 1
            guid_count = len(guid_list)
            if guid_count == recpt_count:
                print(f"GUID = {guid_count} equal to RECPT = {recpt_count}")
            else:
                print("GUID count and RECPT count not match")
                exit(0)
            diss_name = f'MIG-FSQtoEFS-{Tracker_sub_id_value}'
            diss_payload = {
                              "disseminationName": diss_name,
                              "emailTemplateId": "00000000-0000-0000-0000-000000000000",
                              "suppressEmailNotificationFileNoData": False,
                              "feedEntityId": Tracker_feed_id_value,
                              "containerId": Tracker_perm_cont_id_value,
                              "status": 2,
                              "fromAddress": "disseminationnotifications@fefundinfo.com"
                            }
            print(f'Creating Dissemination: {diss_name}')
            diss_found = Search_Diss(diss_name, Tracker_perm_cont_id_value)
            if not diss_found:
                print(f'Diss : {diss_name} not found in {Tracker_perm_cont_id_value}')
                stop_flag = "y"
                # stop_flag = input("Type x to exit")
                if stop_flag == "x":
                    print("do nothing")
                elif stop_flag == "y":
                    print("New Creation")
                    Create_Diss(diss_payload)
                    payload_linking = [{"activated": True, "recipientGuid": guid} for guid in guid_list]
                    print('Linking Recipients')
                    Linking(payload_linking, proj_guid, Tracker_perm_cont_id_value)
            else:
                Get_recpt_count(proj_guid, Tracker_perm_cont_id_value, 0)
                if not linked_recpts:
                    print(f"GUID list is empty for Diss {diss_name} ")
                    print('Linking Recipients')
                    payload_linking = [{"activated": True, "recipientGuid": guid} for guid in guid_list]
                    Linking(payload_linking, proj_guid, Tracker_perm_cont_id_value)
                else:
                    payload_deleting = [{"activated": True, "recipientGuid": guid} for guid in linked_recpts]
                    print('Deleting Recipients')
                    Del_recpts(proj_guid, Tracker_perm_cont_id_value, payload_deleting)
                    print('Linking Recipients')
                    payload_linking = [{"activated": True, "recipientGuid": guid} for guid in guid_list]
                    Linking(payload_linking, proj_guid, Tracker_perm_cont_id_value)
            if flag == 0:
                track_status.value = "RECHECK"
            else:
                track_status.value = "CMPLX"
                flag = 0
            wbt.save(TRACKER)


get_user()
Script_Run()