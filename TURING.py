"""
This program helps with picking the download URL and folder for a specific feed and then downloading the latest file.
The files are then compared and the missing ISIN and headers in the files are listed.
NOTE: The primary Excel file in correct format is required.
"""

import datetime
import os
import subprocess
import time
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import dask.dataframe as dd
import openpyxl
import pandas as pd
import requests
from dask.diagnostics import ProgressBar
from tqdm import tqdm


# Function to read the URL and download path from the Excel file
def read_data_from_excel(file_path):
    data = []
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    efs_url_column = headers.get("EFS")
    hfi_url_column = headers.get("HFI")
    path_column = headers.get("Path")
    if efs_url_column and path_column and hfi_url_column:
        for row in range(2, sheet.max_row + 1):
            efs_url_cell = sheet.cell(row=row, column=efs_url_column)
            hfi_url_cell = sheet.cell(row=row, column=hfi_url_column)
            path_cell = sheet.cell(row=row, column=path_column)
            data.append((efs_url_cell.value, hfi_url_cell.value, path_cell.value))

    return data


def has_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for item in os.listdir(directory_path):
        item_path = os.path.join(directory_path, item)
        if os.path.isdir(item_path):
            return True
    return False


def delete_subdirectories(file_path):
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    for subdirectory in os.listdir(directory_path):
        if os.path.isdir(subdirectory):
            os.system(f'rd /s /q "{subdirectory}"')


def create_subdir(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = {}
    directory_path = os.path.dirname(file_path)
    os.chdir(directory_path)
    path_col = "Path"

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    folder_name = headers.get("Folder Name")
    if folder_name:
        for row in range(2, sheet.max_row + 1):
            folder_cell = sheet.cell(row=row, column=folder_name)
            directory_name = folder_cell.value
            command_run = "mkdir "
            concat_str = " ".join([command_run, directory_name])
            subprocess.run(concat_str, shell=True)
    fetch_path = "dir /ad /b /s"
    fetch_output = subprocess.run(fetch_path, shell=True, capture_output=True)
    # Decode the output and split it into lines
    output_lines = fetch_output.stdout.decode().splitlines()

    # Strip any leading/trailing whitespace from each line
    output_lines = [line.strip() for line in output_lines]

    # Join the lines back together with line breaks
    column_index = None
    for column in sheet.iter_cols(min_row=1, max_row=1):
        if column[0].value == path_col:
            column_index = column[0].column_letter
            break
    # Write the output to the specific column
    if column_index is not None:
        for index, output_line in enumerate(output_lines):
            row_index = index + 2
            sheet[column_index + str(row_index)] = output_line
        wb.save(file_path)
    else:
        print(f"Column '{path_col}' not found.")


# Function to add parameter "&uncompressed=true" to the URL
def add_parameter_to_url(url):
    if url and "&uncompressed=true" not in url:
        if "?" in url:
            url += "&uncompressed=true"
        else:
            url += "?uncompressed=true"
    return url


# Function to download file using requests
def download_file(url, download_path, file_name):
    # Send a GET request to the URL
    response = requests.get(url, stream=True)

    # Get the total file size in bytes
    total_size = int(response.headers.get('content-length', 0))

    # Open the file for writing in binary mode
    with open(download_path + '/' + file_name, 'wb') as file:
        # Create a progress bar using tqdm
        progress_bar = tqdm(total=total_size, unit='B', unit_scale=True)

        # Iterate over the response content in chunks
        for data in response.iter_content(chunk_size=4096):
            # Write the chunk to the file
            file.write(data)

            # Update the progress bar
            progress_bar.update(len(data))

        # Close the progress bar
        progress_bar.close()


def downloadlatest_loop(file_path):
    data = read_data_from_excel(file_path)
    len(data)
    count = len(data)
    print(f"Downloading ", count, "files!")
    efs_file = "ZEFS.csv"
    for efs_url, hfi_url, download_path in data:
        modified_url = add_parameter_to_url(efs_url)
        print(f"Downloading EFS file to", download_path)
        download_file(modified_url, download_path, efs_file)


def download_loop(file_path):
    data = read_data_from_excel(file_path)
    count = len(data) * 2
    print(f"Downloading ", count, "files!")
    efs_file = "ZEFS.csv"
    hfi_file = "AhFI.csv"
    for efs_url, hfi_url, download_path in data:
        modified_url = add_parameter_to_url(efs_url)
        print(f"Downloading EFS file to", download_path)
        download_file(modified_url, download_path, efs_file)
        print(f"Downloading hFI file to", download_path)
        download_file(hfi_url, download_path, hfi_file)


# Main function to automate the task
def automate_task():
    # GUI: Interface to select the input Excel file
    print("Initialising STAGE 1")
    print("---------------------")
    start = time.time()

    file_path = entry_file1.get()

    if file_path:
        try:
            # Read the URL and download path from the Excel file

            if has_subdirectories(file_path):
                if check_latest(file_path):
                    print("hFI files are latest, Download only EFS")
                    # Process each EFS URL and download the LATEST file
                    downloadlatest_loop(file_path)
                else:
                    print('Not Latest! - Refresh folders in progress-')
                    delete_subdirectories(file_path)
                    create_subdir(file_path)
                    download_loop(file_path)
            else:
                print('Fresh Import')
                create_subdir(file_path)
                download_loop(file_path)
        except OSError as e:
            print(f"Error executing .bat file: {e}")
    else:
        print('file path issue')
    if not Flag:
        print(f"Download completed in", time.time()-start)


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def browse_output_file(entry):
    filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[('CSV Files', '*.csv')])
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


def process_tree():
    global Flag
    # Check the state of the checkboxes
    run_automate = var_automate.get()
    run_process = var_process.get()
    Flag = run_process

    # Create and start a new thread for running the UI
    if run_automate:
        automate_task()

    # Run the process_files function in a separate thread
    if run_process:
        process_files()


def check_latest(file_path):
    data_quality = []
    file_name = 'AhFI.csv'
    try:
        df = pd.read_excel(file_path)
    except OSError:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return
    paths = df["Path"].tolist()
    for path in paths:
        if str(path) != "nan":
            if os.path.exists(path):
                # Get the list of files in the directory
                files = os.listdir(path)
                for file in files:
                    if file == file_name:
                        cur_path = os.path.join(path, file_name)
                        modified_timestamp = os.path.getmtime(cur_path)
                        modified_datetime = datetime.datetime.fromtimestamp(modified_timestamp)
                        current_datetime = datetime.datetime.now()
                        time_difference = current_datetime - modified_datetime
                        if time_difference.days < 2:
                            data_quality.append(0)
                            efs_path = os.path.join(path, "ZEFS.csv")
                            if os.path.exists(efs_path):
                                os.remove(efs_path)
                            else:
                                continue
                        else:
                            return False

    for data1 in data_quality:
        if data1 == 0:
            return True
        else:
            return False


def process_files():
    # Get file path
    global OFST020000
    output_data = []
    print("Initialising STAGE 2")
    print("---------------------")
    start_time = time.time()
    file_path = entry_file1.get()

    # Check if file is selected
    if not file_path:
        messagebox.showerror("Error", "Please select a file.")
        return

    # Read Excel file
    try:
        df = pd.read_excel(file_path)
    except OSError:
        messagebox.showerror("Error", "Failed to read the Excel file.")
        return

    # Get the directory path values

    paths = df["Path"].tolist()
    delimiters = df["Delimiter"].tolist()

    # Check if any delimiter is blank and replace with ";"
    for i in range(len(delimiters)):
        if str(delimiters[i]) == "nan":
            delimiters[i] = ";"

    # Iterate over the paths
    for path, delimiter in zip(paths, delimiters):
        # Check if the path is valid
        if os.path.exists(path):
            # Get the list of files in the directory
            files = os.listdir(path)
            # Assuming the first two files in the directory are file 1 and file 2
            if len(files) >= 2:
                file1_path = os.path.join(path, files[1])
                file2_path = os.path.join(path, files[0])

                print("Initialising Data Loading...")
                print(f'File :', file1_path)
                df_efs = dd.read_csv(file1_path, header=0, sep=delimiter, low_memory=False, assume_missing=True,
                                     blocksize=1e9,
                                     skiprows=[1], dtype=str)

                df_hfi = dd.read_csv(file2_path, header=0, sep=delimiter, low_memory=False, assume_missing=True,
                                     blocksize=1e9,
                                     skiprows=[1], dtype=str)

                df_efs = df_efs.compute()

                df_hfi = df_hfi.compute()

                # Get the union of unique headers
                union_headers = pd.unique(df_efs.columns.union(df_hfi.columns))
                # Convert the union_headers to a list if needed
                headers = union_headers.tolist()
                print('--------------')

                # Get headers missing in df_efs
                missing_headers = pd.Series(list(set(headers) - set(df_efs.columns)))
                print('Headers missing in efs:', len(missing_headers))
                head1 = f'Headers missing in {file1_path}:'
                missing_headers_df = pd.DataFrame({head1: missing_headers})
                output_data.append(missing_headers_df)

                # Get ISIN missing in df_efs
                missing_isin = pd.Series(list(set(df_hfi[primary_k].unique()) - set(df_efs[primary_k].unique())))
                print('ISIN missing in efs:', len(missing_isin))
                head2 = f'ISIN missing in {file1_path}'
                missing_isin_df = pd.DataFrame({head2: missing_isin})
                output_data.append(missing_isin_df)
                print("-----------------")

            else:
                messagebox.showwarning("Warning", f"The directory '{path}' does not contain two files.")
        else:
            messagebox.showwarning("Warning", f"The directory path '{path}' does not exist.")
    print("Initialising STAGE 4 - Write data to CSV file")
    output_path = entry_output.get()
    output_data = pd.concat(output_data, axis=1)
    output_data.to_csv(output_path, index=False)
    print("--------------")
    print('Time taken:', time.time() - start_time)
    print("All completed")
    exit(0)


window = tk.Tk()
window.title("Bulk File Comparison")
window.geometry("300x250")
ProgressBar().register()
OFST020000 = 'OFST020000'
Flag = 0

# Create the file selection label and entry
label_file1 = tk.Label(window, text="Select Excel File:")
label_file1.pack()
entry_file1 = tk.Entry(window, width=50)
entry_file1.pack()
button_file1 = tk.Button(window, text="Browse", command=lambda: browse_file(entry_file1))
button_file1.pack()

# Create the output file selection label and entry
label_output = tk.Label(window, text="Output File:")
label_output.pack()
entry_output = tk.Entry(window, width=50)
entry_output.pack()
button_output = tk.Button(window, text="Browse", command=lambda: browse_output_file(entry_output))
button_output.pack()

# Create the checkboxes

var_automate = tk.IntVar()
checkbox_automate = tk.Checkbutton(window, text="DOWNLOAD", variable=var_automate)
checkbox_automate.pack()

var_process = tk.IntVar()
checkbox_process = tk.Checkbutton(window, text="PROCESS FILES", variable=var_process)
checkbox_process.pack()

# Create the process button
button_process = tk.Button(window, text="Process Files", command=process_tree)
button_process.pack()

# Run the main window loop
window.mainloop()
